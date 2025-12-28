#!/usr/bin/env python3 
# -*- coding: utf-8 -*- 
""" 
SAMASTROI SCRAPER (Railway-ready, OFFICIAL) 
- Web-scrapes t.me/s/<channel> pages 
- Builds "cards", enriches with YandexGPT probability (JSON response) 
- Auto-filters by probability threshold 
- Sends cards to a Telegram group with inline action buttons 
- Global decision lock: once any admin clicks ("В работу / Неверно / Привязать") buttons disappear for everyone 
- Training log (JSONL) + decisions & daily aggregates (SQLite) stored on Railway Volume 
- Admin panel: threshold, role management, stats, log, PNG plot, XLSX/PDF reports, KPI dashboard 
- Protection from double poller: single-instance lock on shared volume (prevents 409 getUpdates conflict) 
- "Self-training" (practical): few-shot retrieval + adaptive calibration weights (per-channel bias) stored in DB 
""" 
 
import os 
import shutil 
import re 
import json 
import time 
import uuid 
import sqlite3 
import logging 
import threading 
from datetime import datetime, timezone, date, timedelta 
from typing import Dict, List, Optional, Tuple 
 
import requests 
from bs4 import BeautifulSoup 
 
# matplotlib (PNG plot) 
import matplotlib 
matplotlib.use("Agg") 
import matplotlib.pyplot as plt 
 
# Reports 
from reportlab.pdfgen import canvas 
from reportlab.lib.pagesizes import A4 
from openpyxl import Workbook 
from openpyxl.utils import get_column_letter 
 
logging.basicConfig( 
    format="%(asctime)s | %(levelname)s | %(message)s", 
    level=logging.INFO, 
) 
log = logging.getLogger("samastroi_scraper") 
 
DATA_DIR = os.getenv("DATA_DIR", "/data") 
os.makedirs(DATA_DIR, exist_ok=True) 
 
CARDS_DIR = os.path.join(DATA_DIR, "cards") 
os.makedirs(CARDS_DIR, exist_ok=True) 
 
REPORTS_DIR = os.path.join(DATA_DIR, "reports") 
os.makedirs(REPORTS_DIR, exist_ok=True) 
 
TRAINING_DATASET = os.path.join(DATA_DIR, "training_dataset.jsonl") 
HISTORY_CARDS = os.path.join(DATA_DIR, "history_cards.jsonl") 
SETTINGS_FILE = os.path.join(DATA_DIR, "settings.json") 
DB_PATH = os.path.join(DATA_DIR, "scraper.db") 
LOCK_PATH = os.path.join(DATA_DIR, ".poller.lock") 
 
 
def _seed_config_files() -> None: 
    """Seed /data/groups.txt and /data/keywords.txt from repo files if present. 
 
    Railway volumes persist across deploys. If you keep config files in the repo under ./data, 
    this helper copies them into DATA_DIR on first run so the scraper will load them. 
    """ 
    try: 
        pairs = [ 
            (os.path.join(DATA_DIR, "groups.txt"), ["/app/data/groups.txt", os.path.join(os.getcwd(), "data", "groups.txt")]), 
            (os.path.join(DATA_DIR, "keywords.txt"), ["/app/data/keywords.txt", os.path.join(os.getcwd(), "data", "keywords.txt")]), 
        ] 
        for dst, srcs in pairs: 
            if os.path.isfile(dst): 
                continue 
            src = next((s for s in srcs if s and os.path.isfile(s)), None) 
            if src: 
                os.makedirs(os.path.dirname(dst), exist_ok=True) 
                shutil.copyfile(src, dst) 
                log.info(f"[CFG] seeded {dst} from {src}") 
    except Exception as e: 
        log.warning(f"[CFG] seeding config files failed: {e}") 
 
 
def _ensure_file(path: str, default: str = ""): 
    if not os.path.exists(path): 
        with open(path, "w", encoding="utf-8") as f: 
            f.write(default) 
 
_ensure_file(TRAINING_DATASET) 
_ensure_file(HISTORY_CARDS) 
_ensure_file(SETTINGS_FILE, "{}") 
 
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip() 
TELEGRAM_API_URL = f"https://api.telegram.org/bot{BOT_TOKEN}" if BOT_TOKEN else "" 
 
TARGET_CHAT_ID = int(os.getenv("TARGET_CHAT_ID", "-1003502443229"))  # group/channel for cards 
 
# from user: 
DEFAULT_LEADERSHIP = [5685586625] 
DEFAULT_ADMINS = [272923789, 398960707] 
DEFAULT_MODERATORS = [978125225, 777464055] 
 
YAGPT_API_KEY = os.getenv("YAGPT_API_KEY", "").strip() 
YAGPT_FOLDER_ID = os.getenv("YAGPT_FOLDER_ID", "").strip() 
YAGPT_ENDPOINT = "https://llm.api.cloud.yandex.net/foundationModels/v1/completion" 
YAGPT_MODEL = os.getenv("YAGPT_MODEL", "gpt://{folder_id}/yandexgpt/latest") 
 
SCAN_INTERVAL = int(os.getenv("SCAN_INTERVAL", "300")) 
HTTP_TIMEOUT = int(os.getenv("HTTP_TIMEOUT", "15")) 
MAX_TRAIN_LOG = int(os.getenv("MAX_TRAIN_LOG", "50")) 
TARGET_DATASET_SIZE = int(os.getenv("TARGET_DATASET_SIZE", "5000")) 
DEFAULT_THRESHOLD = int(os.getenv("DEFAULT_THRESHOLD", "0")) 
 
DEFAULT_CHANNELS = [ 
    "tipkhimki", "lobnya", "dolgopacity", "vkhimki", 
    "podslushanovsolnechnogorske", "klingorod", "mspeaks", 
    "pushkino_official", "podmoskow", "trofimovonline", 
    "Tipichnoe_Pushkino", "chp_sergiev_posad", "kraftyou", 
    "kontext_channel", "podslushano_ivanteevka", "pushkino_live", 
    "life_sergiev_posad" 
] 
 
 
# --- Scraper configuration files (optional) --- 
# If CHANNEL_LIST/KEYWORDS env vars are empty, the scraper will load sources/keywords from files. 
# Supported paths: 
#   - /data/groups.txt and /data/keywords.txt (Railway volume) 
#   - /app/data/groups.txt and /app/data/keywords.txt (repo files copied into image) 
GROUPS_FILE = os.getenv("GROUPS_FILE", "").strip() 
KEYWORDS_FILE = os.getenv("KEYWORDS_FILE", "").strip() 
 

# --- Channel blacklist (forced exclusion) ---
BLACKLIST_CHANNELS = {
    "@gsnmo_bou",
    "t.me/gsnmo_bou",
    "gsnmo_bou",
}

def _read_lines_file(path: str) -> List[str]: 
    try: 
        with open(path, "r", encoding="utf-8") as f: 
            return [ln.strip() for ln in f.read().splitlines() if ln.strip() and not ln.strip().startswith("#")] 
    except Exception: 
        return [] 
 
def _find_first_existing(paths: List[str]) -> Optional[str]: 
    for p in paths: 
        try: 
            if p and os.path.isfile(p): 
                return p 
        except Exception: 
            pass 
    return None 
 
def _normalize_source(s: str) -> Optional[str]: 
    """Normalize a channel/group source token to a public username usable in https://t.me/s/<username>. 
 
    Accepts: 
      - @username 
      - username 
      - https://t.me/username or https://t.me/s/username 
    Skips: 
      - numeric chat ids (-100..., 12345) because web preview can't use them 
      - invite links (joinchat/+...) 
    """ 
    s = (s or "").strip() 
    if not s: 
        return None 
 
    # drop comments in files 
    if s.startswith("#") or s.startswith("//"): 
        return None 
 
    s = s.replace("https://", "").replace("http://", "") 
 
    # skip invite links (not web-scrapable) 
    low = s.lower() 
    if "joinchat" in low or "/+" in low or low.startswith("+"): 
        return None 
 
    s = s.replace("t.me/s/", "t.me/").replace("t.me/", "") 
    s = s.lstrip("@").strip().strip("/") 
    if not s: 
        return None 
 
    # keep only username part before any query params 
    s = s.split("?")[0].split("#")[0].strip() 
 
    # skip numeric ids (web preview requires username) 
    if re.fullmatch(r"-?\d+", s): 
        return None 
 
    return s.lower() 
 
def load_channel_list() -> List[str]: 
    env = (os.getenv("CHANNEL_LIST", "") or "").strip() 
    if env: 
        items = [x.strip() for x in env.split(",") if x.strip()] 
        res = [] 
        for it in items: 
            n = _normalize_source(it) 
            if n: 
                res.append(n) 
        return res 
 
    candidate_files = [] 
    if GROUPS_FILE: 
        candidate_files.append(GROUPS_FILE) 
    candidate_files += [os.path.join(DATA_DIR, "groups.txt"), "/app/data/groups.txt", os.path.join(os.getcwd(), "data", "groups.txt")] 
    fp = _find_first_existing(candidate_files) 
    if fp: 
        raw = _read_lines_file(fp) 
        res = [] 
        for it in raw: 
            n = _normalize_source(it) 
            if n: 
                res.append(n) 
        if res: 
            log.info(f"[CFG] channels loaded from {fp}: {len(res)}") 
            return res 
 
    log.info("[CFG] channels fallback to DEFAULT_CHANNELS") 
    return list(DEFAULT_CHANNELS) 
 
def load_keywords_list() -> List[str]: 
    env = (os.getenv("KEYWORDS", "") or "").strip() 
    if env: 
        # allow comma or newline separated 
        parts = [] 
        for chunk in env.split(","): 
            chunk = chunk.strip() 
            if chunk: 
                parts.append(chunk) 
        return parts 
 
    candidate_files = [] 
    if KEYWORDS_FILE: 
        candidate_files.append(KEYWORDS_FILE) 
    candidate_files += [os.path.join(DATA_DIR, "keywords.txt"), "/app/data/keywords.txt", os.path.join(os.getcwd(), "data", "keywords.txt")] 
    fp = _find_first_existing(candidate_files) 
    if fp: 
        res = _read_lines_file(fp) 
        if res: 
            log.info(f"[CFG] keywords loaded from {fp}: {len(res)}") 
            return res 
 
    # Fallback keywords (kept minimal; prefer keywords.txt) 
    return [ 
        "стройка", "строительство", "самострой", "самосострой", "незаконная стройка", 
        "котлован", "фундамент", "кран", "экскаватор", "застройщик", "ОНзС", "онзс", "объект незаконного строительства" 
    ] 
 
CHANNEL_LIST = load_channel_list() 

def normalize_channel(ch: str) -> str:
    ch = (ch or "").strip()
    if not ch:
        return ch
    if ch.startswith("https://t.me/"):
        ch = "@" + ch.split("https://t.me/", 1)[1]
    if ch.startswith("t.me/"):
        ch = "@" + ch.split("t.me/", 1)[1]
    if not ch.startswith("@"):
        ch = "@" + ch
    return ch.lower()

# Apply blacklist (defensive: works even if ENV is misconfigured)
_before = len(CHANNEL_LIST)
_bl = {normalize_channel(b) for b in BLACKLIST_CHANNELS}
CHANNEL_LIST = [ch for ch in CHANNEL_LIST if normalize_channel(ch) not in _bl]
_after = len(CHANNEL_LIST)
log.info(f"[CFG] Blacklist applied. Channels before={_before}, after={_after}")

KEYWORDS = load_keywords_list() 
 
# Extra high-signal patterns (work even without keywords)
CADASTRE_RE = re.compile(r"\b\d{2}:\d{2}:\d{6,8}:\d+\b")
COORD_RE = re.compile(r"\b\d{2}\.\d{3,}\s*,\s*\d{2}\.\d{3,}\b")
ADDRESS_RE = re.compile(r'\b(ул\.?|улица|проспект|пр-т\.?|площадь|пл\.?|переулок|пер\.?|шоссе|ш\.?)\s+([\w\s-]+?)\s+(д\.?|дом)\s+(\d+([\w\/]*))', re.IGNORECASE)

# CHANNEL_LIST is loaded via load_channel_list() above
# KEYWORDS are loaded via load_keywords_list() above 
KEYWORDS_LOWER = [k.lower() for k in KEYWORDS] 
 
 
def db() -> sqlite3.Connection: 
    conn = sqlite3.connect(DB_PATH, timeout=30, isolation_level=None) 
    conn.execute("PRAGMA journal_mode=WAL;") 
    conn.execute("PRAGMA synchronous=NORMAL;") 
    return conn 
 
def init_db(): 
    conn = db() 
 
    conn.execute(""" 
        CREATE TABLE IF NOT EXISTS seen_posts ( 
            channel TEXT NOT NULL, 
            post_id TEXT NOT NULL, 
            first_seen_ts INTEGER NOT NULL, 
            PRIMARY KEY (channel, post_id) 
        ); 
    """) 
 
    conn.execute(""" 
        CREATE TABLE IF NOT EXISTS card_decisions ( 
            card_id TEXT PRIMARY KEY, 
            decision TEXT NOT NULL, 
            decided_by INTEGER NOT NULL, 
            decided_ts INTEGER NOT NULL 
        ); 
    """) 
 
    conn.execute("""
        CREATE TABLE IF NOT EXISTS card_status (
            card_id TEXT PRIMARY KEY,
            onzs_category INTEGER,
            status TEXT,
            comment TEXT,
            last_updated_ts INTEGER NOT NULL,
            last_updated_by INTEGER NOT NULL,
            FOREIGN KEY (card_id) REFERENCES card_decisions (card_id)
        );
    """)
 
    conn.execute(""" 
        CREATE TABLE IF NOT EXISTS train_daily ( 
            day TEXT PRIMARY KEY, 
            total INTEGER NOT NULL, 
            work INTEGER NOT NULL, 
            wrong INTEGER NOT NULL, 
            attach INTEGER NOT NULL 
        ); 
    """) 
 
    conn.execute(""" 
        CREATE TABLE IF NOT EXISTS user_roles ( 
            user_id INTEGER PRIMARY KEY, 
            role TEXT NOT NULL 
        ); 
    """) 
 
    conn.execute(""" 
        CREATE TABLE IF NOT EXISTS model_params ( 
            key TEXT PRIMARY KEY, 
            value_json TEXT NOT NULL 
        ); 
    """) 
 
    
    conn.execute("""
        CREATE TABLE IF NOT EXISTS rgis_cache (
            cadastral TEXT PRIMARY KEY,
            result_json TEXT NOT NULL,
            fetched_ts INTEGER NOT NULL
        );
    """)
    # seed roles if empty 
    cnt = int(conn.execute("SELECT COUNT(*) FROM user_roles;").fetchone()[0] or 0) 
    if cnt == 0: 
        for uid in DEFAULT_LEADERSHIP: 
            conn.execute("INSERT OR REPLACE INTO user_roles(user_id, role) VALUES (?, ?);", (int(uid), "leadership")) 
        for uid in DEFAULT_ADMINS: 
            conn.execute("INSERT OR REPLACE INTO user_roles(user_id, role) VALUES (?, ?);", (int(uid), "admin")) 
        for uid in DEFAULT_MODERATORS: 
            conn.execute("INSERT OR REPLACE INTO user_roles(user_id, role) VALUES (?, ?);", (int(uid), "moderator")) 
 
    conn.execute( 
        "INSERT OR IGNORE INTO model_params(key, value_json) VALUES (?, ?);", 
        ("threshold", json.dumps({"value": DEFAULT_THRESHOLD}, ensure_ascii=False)) 
    ) 
    # weights: per-channel bias in probability points ([-25..25]) + label weights for aggregation 
    conn.execute( 
        "INSERT OR IGNORE INTO model_params(key, value_json) VALUES (?, ?);", 
        ("weights", json.dumps({"channels": {}, "label_weights": {"work": 1.0, "wrong": 1.0, "attach": 1.0}}, ensure_ascii=False)) 
    ) 
 
    conn.commit() 
    conn.close() 
 
init_db() 
 
 
def load_json(path: str, default): 
    try: 
        with open(path, "r", encoding="utf-8") as f: 
            return json.load(f) 
    except Exception: 
        return default 
 
def save_json(path: str, obj): 
    tmp = path + ".tmp" 
    with open(tmp, "w", encoding="utf-8") as f: 
        json.dump(obj, f, ensure_ascii=False, indent=2) 
    os.replace(tmp, path) 
 
def load_settings() -> Dict: 
    return load_json(SETTINGS_FILE, {}) 
 
def save_settings(s: Dict): 
    save_json(SETTINGS_FILE, s) 
 
def get_prob_threshold() -> int: 
    s = load_settings() 
    try: 
        v = int(s.get("prob_threshold", DEFAULT_THRESHOLD)) 
        return max(0, min(100, v)) 
    except Exception: 
        return DEFAULT_THRESHOLD 
 
def set_prob_threshold(v: int): 
    v = max(0, min(100, int(v))) 
    s = load_settings() 
    s["prob_threshold"] = v 
    save_settings(s) 
 
def get_update_offset() -> int: 
    s = load_settings() 
    try: 
        return int(s.get("update_offset", 0)) 
    except Exception: 
        return 0 
 
def set_update_offset(v: int): 
    s = load_settings() 
    s["update_offset"] = int(v) 
    save_settings(s) 
 
def get_role(user_id: int) -> Optional[str]: 
    conn = db() 
    row = conn.execute("SELECT role FROM user_roles WHERE user_id=?;", (int(user_id),)).fetchone() 
    conn.close() 
    return row[0] if row else None 
 
def is_admin(user_id: int) -> bool: 
    return get_role(int(user_id)) == "admin" 
 
def is_moderator(user_id: int) -> bool: 
    return get_role(int(user_id)) == "moderator" 
 
def is_leadership(user_id: int) -> bool: 
    return get_role(int(user_id)) == "leadership" 
 
def list_users_by_role(role: str) -> List[int]: 
    conn = db() 
    rows = conn.execute("SELECT user_id FROM user_roles WHERE role=? ORDER BY user_id;", (role,)).fetchall() 
    conn.close() 
    return [int(r[0]) for r in rows] 
 
def upsert_role(user_id: int, role: str) -> None: 
    conn = db() 
    conn.execute("INSERT OR REPLACE INTO user_roles(user_id, role) VALUES (?, ?);", (int(user_id), role)) 
    conn.commit() 
    conn.close() 
 
def remove_role(user_id: int) -> None: 
    conn = db() 
    conn.execute("DELETE FROM user_roles WHERE user_id=?;", (int(user_id),)) 
    conn.commit() 
    conn.close() 
 
def add_admin(user_id: int): upsert_role(int(user_id), "admin") 
def add_moderator(user_id: int): upsert_role(int(user_id), "moderator") 
def add_leadership(user_id: int): upsert_role(int(user_id), "leadership") 
 
def remove_admin(user_id: int): remove_role(int(user_id)) 
def remove_moderator(user_id: int): remove_role(int(user_id)) 
def remove_leadership(user_id: int): remove_role(int(user_id)) 
 
def acquire_lock_or_exit() -> None: 
    """ 
    Prevent multiple instances from running getUpdates poller on Railway volume. 
    """ 
    try: 
        fd = os.open(LOCK_PATH, os.O_CREAT | os.O_EXCL | os.O_WRONLY) 
        with os.fdopen(fd, "w", encoding="utf-8") as f: 
            f.write(str(os.getpid())) 
        log.info(f"Lock acquired: {LOCK_PATH}") 
    except FileExistsError: 
        log.error("Another instance holds poller lock. Exiting (prevents 409 getUpdates conflict).") 
        raise SystemExit(0) 
 
def release_lock(): 
    try: 
        if os.path.exists(LOCK_PATH): 
            os.remove(LOCK_PATH) 
            log.info("Lock released.") 
    except Exception: 
        pass 
 
def now_ts() -> int: 
    return int(time.time()) 
 
def append_jsonl(path: str, obj: Dict): 
    with open(path, "a", encoding="utf-8") as f: 
        f.write(json.dumps(obj, ensure_ascii=False) + "\n") 
 

def rgis_cache_get(cadastral: str, ttl_hours: int = None) -> Optional[Dict]:
    cadastral = (cadastral or "").strip()
    if not cadastral:
        return None
    if ttl_hours is None:
        ttl_hours = RGIS_CACHE_TTL_HOURS
    conn = db()
    try:
        row = conn.execute(
            "SELECT result_json, fetched_ts FROM rgis_cache WHERE cadastral=?;",
            (cadastral,)
        ).fetchone()
        if not row:
            return None
        result_json, fetched_ts = row
        if now_ts() - int(fetched_ts) > int(ttl_hours) * 3600:
            return None
        obj = json.loads(result_json)
        if isinstance(obj, dict):
            obj["cache"] = True
        return obj
    except Exception:
        return None
    finally:
        conn.close()

def rgis_cache_put(cadastral: str, result: Dict) -> None:
    cadastral = (cadastral or "").strip()
    if not cadastral:
        return
    conn = db()
    try:
        conn.execute(
            "INSERT OR REPLACE INTO rgis_cache(cadastral, result_json, fetched_ts) VALUES (?,?,?);",
            (cadastral, json.dumps(result, ensure_ascii=False), now_ts())
        )
    finally:
        conn.close()

def rgis_lookup_playwright(cadastral: str, timeout_ms: int = 25000) -> Dict:
    """
    Реальный парсинг RGIS (planning) через Playwright (Chromium):
      - открывает https://rgis.mosreg.ru/v3/#/?tab=planning
      - вводит кадастровый номер и запускает поиск
      - пытается считать текст боковой панели/карточки результата
    """
    cadastral = (cadastral or "").strip()
    if not cadastral:
        return {"ok": False, "error": "empty cadastral"}

    cached = rgis_cache_get(cadastral)
    if cached:
        return cached

    url = "https://rgis.mosreg.ru/v3/#/?tab=planning"
    result: Dict = {
        "ok": False,
        "cadastral": cadastral,
        "url": url,
        "cache": False,
        "fetched_ts": now_ts(),
    }

    try:
        from playwright.sync_api import sync_playwright
    except Exception as e:
        result["error"] = f"Playwright is not installed/available: {e}"
        rgis_cache_put(cadastral, result)
        return result

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"],
        )
        context = browser.new_context(viewport={"width": 1400, "height": 900})
        page = context.new_page()
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=timeout_ms)
            page.wait_for_load_state("networkidle", timeout=timeout_ms)

            for sel in [
                "button:has-text('Принять')",
                "button:has-text('Согласен')",
                "button:has-text('OK')",
                "button:has-text('Понятно')",
            ]:
                try:
                    loc = page.locator(sel).first
                    if loc.is_visible(timeout=800):
                        loc.click(timeout=800)
                        break
                except Exception:
                    pass

            search_locators = [
                "input[placeholder*='Поиск' i]",
                "input[placeholder*='кадастр' i]",
                "input[placeholder*='кадастров' i]",
                "input[type='search']",
                "input",
            ]
            search_input = None
            for sel in search_locators:
                try:
                    loc = page.locator(sel).first
                    if loc.is_visible(timeout=1500):
                        loc.click(timeout=800)
                        search_input = loc
                        break
                except Exception:
                    continue

            if not search_input:
                result["error"] = "search input not found"
                rgis_cache_put(cadastral, result)
                return result

            try:
                search_input.fill("")
                search_input.type(cadastral, delay=20)
            except Exception:
                try:
                    search_input.fill(cadastral)
                except Exception as e:
                    result["error"] = f"cannot type cadastral: {e}"
                    rgis_cache_put(cadastral, result)
                    return result

            clicked = False
            for sel in [
                "button:has-text('Найти')",
                "button:has-text('Поиск')",
                "button[aria-label*='Поиск' i]",
                "button[title*='Поиск' i]",
            ]:
                try:
                    loc = page.locator(sel).first
                    if loc.is_visible(timeout=1200):
                        loc.click(timeout=1200)
                        clicked = True
                        break
                except Exception:
                    continue
            if not clicked:
                try:
                    page.keyboard.press("Enter")
                except Exception:
                    pass

            page.wait_for_timeout(1200)
            page.wait_for_load_state("networkidle", timeout=timeout_ms)

            panel_text = ""
            used_sel = None
            for sel in [
                "[class*='sidebar' i]",
                "[class*='drawer' i]",
                "[class*='panel' i]",
                "main",
                "body",
            ]:
                try:
                    loc = page.locator(sel).first
                    if loc.is_visible(timeout=1500):
                        txt = (loc.inner_text(timeout=2000) or "").strip()
                        if len(txt) > 120:
                            panel_text = txt
                            used_sel = sel
                            break
                except Exception:
                    continue

            if panel_text:
                result["ok"] = True
                result["panel_selector"] = used_sel
                result["panel_text"] = panel_text[:12000]
            else:
                result["error"] = "no meaningful panel text after search"

            rgis_cache_put(cadastral, result)
            return result
        except Exception as e:
            result["error"] = f"rgis_lookup_playwright exception: {e}"
            rgis_cache_put(cadastral, result)
            return result
        finally:
            try:
                context.close()
            except Exception:
                pass
            try:
                browser.close()
            except Exception:
                pass

def add_onzs_trace(card: Dict, step: str, data: Dict) -> None:
    """Append ONZS trace step into card['onzs_trace'] preserving order."""
    try:
        card.setdefault("onzs_trace", [])
        card["onzs_trace"].append({
            "ts": now_ts(),
            "step": step,
            "data": data or {}
        })
    except Exception:
        pass

def normalize_text(text: str) -> str: 
    if not isinstance(text, str): 
        return "" 
    text = text.replace("\n", " ").replace("\t", " ") 
    text = " ".join(text.split()) 
    return text.strip() 
 
def detect_keywords(text: str) -> List[str]: 
    low = (text or "").lower() 
    hits = [kw for kw in KEYWORDS_LOWER if kw and kw in low] 
 
    # tolerate common typos/variants around "самострой" 
    if any(x in low for x in ("самостро", "самосостро", "самоcтро", "самос т ро")): 
        if "самострой" not in hits: 
            hits.append("самострой") 
 
    # extra high-signal patterns (even if keywords list is short) 
    if CADASTRE_RE.search(text or ""): 
        hits.append("кадастровый номер") 
    if COORD_RE.search(text or ""): 
        hits.append("координаты") 
 
    # de-dup while preserving order 
    out = [] 
    seen = set() 
    for h in hits: 
        if h in seen: 
            continue 
        seen.add(h) 
        out.append(h) 
    return out 
 
def parse_tg_datetime(dt_str: str) -> int: 
    """ 
    Telegram embeds: <time datetime="2025-12-14T16:27:14+00:00"> 
    Convert to unix ts. 
    """ 
    if not dt_str: 
        return now_ts() 
    try: 
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00")) 
        if dt.tzinfo is None: 
            dt = dt.replace(tzinfo=timezone.utc) 
        return int(dt.timestamp()) 
    except Exception: 
        return now_ts() 
 
 
def _get_model_param(key: str, default_obj: Dict) -> Dict: 
    conn = db() 
    row = conn.execute("SELECT value_json FROM model_params WHERE key=?;", (key,)).fetchone() 
    conn.close() 
    if not row: 
        return default_obj 
    try: 
        return json.loads(row[0]) 
    except Exception: 
        return default_obj 
 
def _set_model_param(key: str, obj: Dict) -> None: 
    conn = db() 
    conn.execute("INSERT OR REPLACE INTO model_params(key, value_json) VALUES (?, ?);", (key, json.dumps(obj, ensure_ascii=False))) 
    conn.commit() 
    conn.close() 
 
def get_channel_bias(channel: str) -> float: 
    w = _get_model_param("weights", {"channels": {}, "label_weights": {}}) 
    try: 
        return float(w.get("channels", {}).get(channel, 0.0)) 
    except Exception: 
        return 0.0 
 
def update_channel_bias(channel: str, label: str) -> None: 
    """ 
    Adaptive calibration: 
    - If admins often mark channel posts as "work/attach" -> increase bias (more aggressive) 
    - If often "wrong" -> decrease bias 
    Stored as +/- probability points. 
    """ 
    w = _get_model_param("weights", {"channels": {}, "label_weights": {"work": 1.0, "wrong": 1.0, "attach": 1.0}}) 
    ch = w.setdefault("channels", {}) 
    cur = float(ch.get(channel, 0.0) or 0.0) 
 
    step = 1.5  # points per decision 
    if label == "work": 
        cur += step 
    elif label == "wrong": 
        cur -= step 
 
    cur = max(-25.0, min(25.0, cur)) 
    ch[channel] = round(cur, 2) 
    _set_model_param("weights", w) 
 
 
def mark_seen(channel: str, post_id: str, ts: int) -> bool: 
    conn = db() 
    try: 
        conn.execute("INSERT OR IGNORE INTO seen_posts(channel, post_id, first_seen_ts) VALUES(?,?,?)", (channel, post_id, ts)) 
        changed = conn.execute("SELECT changes()").fetchone()[0] == 1 
        return changed 
    finally: 
        conn.close() 
 
def decision_exists(card_id: str) -> Optional[Tuple[str, int, int]]: 
    conn = db() 
    try: 
        row = conn.execute("SELECT decision, decided_by, decided_ts FROM card_decisions WHERE card_id=?", (card_id,)).fetchone() 
        return row if row else None 
    finally: 
        conn.close() 
 
def set_decision(card_id: str, decision: str, user_id: int) -> bool: 
    """ 
    Idempotent global decision: only first admin click is accepted. 
    """ 
    conn = db() 
    try: 
        conn.execute("BEGIN IMMEDIATE;") 
        row = conn.execute("SELECT card_id FROM card_decisions WHERE card_id=?", (card_id,)).fetchone() 
        if row: 
            conn.execute("COMMIT;") 
            return False 
        conn.execute( 
            "INSERT INTO card_decisions(card_id, decision, decided_by, decided_ts) VALUES(?,?,?,?)", 
            (card_id, decision, int(user_id), now_ts()), 
        ) 
        conn.execute("COMMIT;") 
        return True 
    except Exception: 
        conn.execute("ROLLBACK;") 
        raise 
    finally: 
        conn.close() 
 
def update_train_daily(label: str): 
    d = date.today().isoformat() 
    conn = db() 
    try: 
        row = conn.execute("SELECT total, work, wrong, attach FROM train_daily WHERE day=?", (d,)).fetchone() 
        if row: 
            total, work, wrong, attach = row 
        else: 
            total = work = wrong = attach = 0 
        total += 1 
        if label == "work": 
            work += 1 
        elif label == "wrong": 
            wrong += 1        conn.execute( 
            "INSERT OR REPLACE INTO train_daily(day,total,work,wrong,attach) VALUES(?,?,?,?,?)", 
            (d, total, work, wrong, attach), 
        ) 
    finally: 
        conn.close() 
 
def log_training_event(card_id: str, label: str, text: str, channel: str, admin_id: int): 
    rec = { 
        "timestamp": now_ts(), 
        "card_id": card_id, 
        "label": label, 
        "admin_id": int(admin_id), 
        "channel": channel, 
        "text": (text or "")[:5000], 
    } 
    append_jsonl(TRAINING_DATASET, rec) 
    update_train_daily(label) 
    update_channel_bias(channel, label) 
 
def compute_training_stats() -> Dict: 
    """Возвращает агрегированную статистику обучения. 
 
    Режимы: 
    - STATS_MODE=override (по умолчанию): показывает фиксированные/настроечные цифры (удобно для демонстрации). 
    - STATS_MODE=auto: считает по базе (train_daily) и датасету. 
    """ 
    mode = (os.getenv("STATS_MODE") or "auto").strip().lower() 
 
    if mode in ("override", "fixed", "demo", "1", "true", "yes"): 
        total = int(os.getenv("STATS_TOTAL", "3246")) 
        work = int(os.getenv("STATS_IN_WORK", "201")) 
        wrong = int(os.getenv("STATS_WRONG", "3045"))        target = int(os.getenv("STATS_TARGET", "5000")) 
 
        # прогресс к цели — от total/target 
        prog = 0.0 if target <= 0 else (total / target) * 100.0 
        # условная уверенность — доля "В работу" от общего объёма 
        conf = 0.0 if total <= 0 else (work / total) * 100.0 
 
        # Форматируем в стиле RU (запятая) 
        prog_s = f"{prog:.1f}".replace(".", ",") 
        conf_s = f"{conf:.1f}".replace(".", ",") 
 
        last_str = (os.getenv("STATS_LAST_EVENT") or "25.12.2025 09:06").strip() 
        last_ts = None 
        try: 
            dt = datetime.strptime(last_str, "%d.%m.%Y %H:%M") 
            # timestamp в локальном времени контейнера 
            last_ts = int(time.mktime(dt.timetuple())) 
        except Exception: 
            pass 
 
        return { 
            "total": total, 
            "work": work, 
            "wrong": wrong, 
                        "progress": prog_s, 
            "confidence": conf_s, 
            "last_ts": last_ts, 
            "last_str": last_str, 
            "target": target, 
        } 
 
    # === AUTO режим (как было раньше) === 
    conn = db() 
    rows = conn.execute("SELECT total, work, wrong, attach FROM train_daily").fetchall() 
    conn.close() 
 
    total = sum(r[0] for r in rows) if rows else 0 
    work = sum(r[1] for r in rows) if rows else 0 
    wrong = sum(r[2] for r in rows) if rows else 0    last_ts = None 
    try: 
        with open(TRAINING_DATASET, "rb") as f: 
            f.seek(0, os.SEEK_END) 
            size = f.tell() 
            if size > 0: 
                f.seek(max(0, size - 8192), os.SEEK_SET) 
                chunk = f.read().decode("utf-8", errors="ignore") 
                lines = [ln for ln in chunk.splitlines() if ln.strip()] 
                for ln in reversed(lines): 
                    try: 
                        obj = json.loads(ln) 
                        ts = obj.get("timestamp") 
                        if isinstance(ts, int): 
                            last_ts = ts 
                            break 
                    except Exception: 
                        continue 
    except Exception: 
        pass 
 
    prog = 0.0 if TARGET_DATASET_SIZE <= 0 else min(1.0, total / TARGET_DATASET_SIZE) * 100.0
    conf = 0.0 if total <= 0 else (work / total) * 100.0
    return {
        "total": total,
        "work": work,
        "wrong": wrong,
                "progress": round(prog, 2),
        "confidence": round(conf, 2),
        "last_ts": last_ts,
        "target": TARGET_DATASET_SIZE,
    }
 
def tail_training_log(limit: int = MAX_TRAIN_LOG) -> List[Dict]: 
    if not os.path.exists(TRAINING_DATASET): 
        return [] 
    try: 
        with open(TRAINING_DATASET, "r", encoding="utf-8") as f: 
            lines = f.readlines() 
        out = [] 
        for ln in lines[-limit:]: 
            ln = ln.strip() 
            if not ln: 
                continue 
            try: 
                out.append(json.loads(ln)) 
            except Exception: 
                pass 
        return out 
    except Exception: 
        return [] 
 
def sparkline(values: List[int]) -> str: 
    if not values: 
        return "—" 
    blocks = "▁▂▃▄▅▆▇█" 
    mn, mx = min(values), max(values) 
    if mx == mn: 
        return blocks[0] * len(values) 
    out = [] 
    for v in values: 
        idx = int((v - mn) * (len(blocks) - 1) / (mx - mn)) 
        out.append(blocks[idx]) 
    return "".join(out) 
 
def training_plot_text(days: int = 14) -> str: 
    rows = _fetch_train_daily_last(days) 
    if not rows: 
        return "📊 График роста обучения: данных пока нет." 
    labels = [r[0][5:] for r in rows] 
    totals = [int(r[1]) for r in rows] 
    return "📊 График роста обучения (событий в день):\n" + sparkline(totals) + "\n" + " | ".join(f"{labels[i]}:{totals[i]}" for i in range(len(labels))) 
 
 
def select_few_shot_examples(text: str, k: int = 3) -> List[Dict]: 
    """ 
    Retrieves recent labeled examples by keyword overlap for few-shot calibration. 
    """ 
    keys = set(detect_keywords((text or "").lower())) 
    if not keys: 
        return [] 
    events = tail_training_log(limit=250) 
    scored = [] 
    for e in events: 
        t = (e.get("text") or "").lower() 
        if not t: 
            continue 
        ekeys = set(detect_keywords(t)) 
        score = len(keys & ekeys) 
        if score > 0: 
            scored.append((score, e)) 
    scored.sort(key=lambda x: x[0], reverse=True) 
    return [e for _, e in scored[:k]] 
 

def call_yandex_gpt_json(card: Dict) -> Tuple[Optional[Dict], Optional[str]]:
    """
    Calls YandexGPT and expects STRICT JSON in response.

    Expected schema (keys may be partial, we will normalize):
    {
      "probability": number (0..100),
      "comment": string,
      "onzs_category_id": number (1..12),
      "onzs_category_name": string,
      "onzs_confidence": number (0..1),
      "risk_level": "low|medium|high",
      "risk_score": number (0..100),
      "recommendation": string,
      "justification": string
    }
    """
    if not YAGPT_API_KEY or not YAGPT_FOLDER_ID:
        return None, "YandexGPT API Key or Folder ID is not configured."

    model_uri = YAGPT_MODEL.format(folder_id=YAGPT_FOLDER_ID)

    # Provide the model a compact, high-signal payload (no huge histories)
    payload_for_ai = {
        "card_id": card.get("card_id"),
        "channel": card.get("channel"),
        "post_id": card.get("post_id"),
        "timestamp": card.get("timestamp"),
        "text": (card.get("text") or "")[:5000],
        "keywords": card.get("keywords") or [],
        "geo": card.get("geo") or card.get("geo_info") or {},
        "object": card.get("object") or {},
        "signals": (card.get("text_profile") or {}).get("signals") if card.get("text_profile") else (card.get("signals") or {}),
        "municipality_hint": ((card.get("geo") or {}).get("municipality") or (card.get("geo_info") or {}).get("municipality_hint")),
        "cadastral_number": ((card.get("geo") or {}).get("cadastral_number") or (card.get("geo_info") or {}).get("cadastral_number")),
        "coordinates": ((card.get("geo") or {}).get("coordinates") or (card.get("geo_info") or {}).get("coordinates")),
        "address": ((card.get("geo") or {}).get("address") or (card.get("geo_info") or {}).get("address")),
        "onzs_candidates": [{"id": k, "name": v.get("name")} for k, v in ONZS_CATEGORIES.items()],
    }

    prompt = (
        "Ты — эксперт органа государственного строительного надзора Московской области.\n"
        "Проанализируй карточку и оцени:\n"
        "1) вероятность самовольного строительства probability (0..100)\n"
        "2) риск: risk_score (0..100) и risk_level (low/medium/high)\n"
        "3) категорию ОНзС: onzs_category_id (1..12) и onzs_category_name\n"
        "4) рекомендацию: recommendation (например: 'Рекомендуется выездная проверка', "
        "'Рекомендуется мониторинг', 'Можно архивировать')\n"
        "5) обоснование: justification (1–3 предложения, строго по данным)\n\n"
        "Верни СТРОГО JSON следующего вида (без текста вокруг):\n"
        "{\n"
        '  "probability": 0,\n'
        '  "comment": "",\n'
        '  "onzs_category_id": 0,\n'
        '  "onzs_category_name": "",\n'
        '  "onzs_confidence": 0.0,\n'
        '  "risk_level": "low",\n'
        '  "risk_score": 0,\n'
        '  "recommendation": "",\n'
        '  "justification": ""\n'
        "}\n\n"
        "Данные карточки (JSON):\n"
        + json.dumps(payload_for_ai, ensure_ascii=False)
    )

    body = {
        "modelUri": model_uri,
        "completionOptions": {"stream": False, "temperature": 0.1, "maxTokens": 450},
        "messages": [{"role": "user", "text": prompt}],
    }
    headers = {
        "Authorization": f"Api-Key {YAGPT_API_KEY}",
        "x-folder-id": YAGPT_FOLDER_ID,
        "Content-Type": "application/json",
    }

    try:
        resp = requests.post(YAGPT_ENDPOINT, headers=headers, json=body, timeout=25)
        if resp.status_code != 200:
            return None, f"API Request Failed with status {resp.status_code}: {resp.text}"
        data = resp.json()
    except requests.exceptions.RequestException as e:
        log.error(f"YandexGPT request error: {e}")
        return None, f"API Request Failed: {e}"
    except json.JSONDecodeError:
        log.error(f"YandexGPT JSON decode error. Response: {getattr(resp,'text','')}")
        return None, "Failed to decode API response."

    try:
        text_out = data["result"]["alternatives"][0]["message"]["text"]
    except (KeyError, IndexError) as e:
        log.error(f"YandexGPT response parse error: {e}; data={data}")
        return None, "Unexpected API response format."

    out = (text_out or "").strip()
    # try to extract JSON object
    if not out.startswith("{"):
        s = out.find("{")
        e = out.rfind("}")
        if s != -1 and e != -1 and e > s:
            out = out[s:e + 1]

    try:
        return json.loads(out), None
    except json.JSONDecodeError as e:
        log.error(f"YandexGPT JSON parse error: {e}; text={text_out[:300]}")
        return None, "Failed to parse JSON from AI response."


def enrich_card_with_yagpt(card: Dict) -> None:
    """
    Enrich card with AI decision:
    - probability (0..100) + comment
    - ONZS category + justification
    - risk_score + risk_level + recommendation
    Also performs hybrid merge with rule-based risk score.
    """
    t = (card.get("text") or "").strip()
    if not t:
        return

    # baseline risk from rules (for hybrid)
    obj = card.get("object") or {}
    signals = (card.get("text_profile") or {}).get("signals") if card.get("text_profile") else card.get("signals") or {}
    geo = card.get("geo") or card.get("geo_info") or {}
    has_cad = bool(geo.get("cadastral_number"))
    has_coords = bool(geo.get("coordinates"))
    base_score, base_level = rule_based_risk_score(
        obj_type=obj.get("type"),
        stage=obj.get("stage"),
        signals=signals or {},
        has_cadastre=has_cad,
        has_coords=has_coords,
    )

    card.setdefault("ai", {})
    card["ai"]["risk_score_rule"] = base_score
    card["ai"]["risk_level_rule"] = base_level
    card["ai"]["model"] = "yandexgpt"

    res, err = call_yandex_gpt_json(card)
    if err:
        card["ai"]["error"] = err
        return
    if not res:
        card["ai"]["error"] = "AI returned no result."
        return

    # Normalize fields
    def _num(x, default=None):
        try:
            if x is None:
                return default
            return float(x)
        except Exception:
            return default

    prob = _num(res.get("probability"), None)
    if prob is not None:
        prob = max(0.0, min(100.0, prob))

    risk_score_ai = _num(res.get("risk_score"), None)
    if risk_score_ai is not None:
        risk_score_ai = max(0.0, min(100.0, risk_score_ai))

    risk_level_ai = (res.get("risk_level") or "").strip().lower()
    if risk_level_ai not in ("low", "medium", "high"):
        risk_level_ai = None

    # Hybrid: weighted blend, AI can adjust but not fully override
    # final = 0.65 rules + 0.35 AI (if available)
    if risk_score_ai is None:
        final_score = base_score
    else:
        final_score = int(round(base_score * 0.65 + risk_score_ai * 0.35))

    if final_score >= 70:
        final_level = "high"
    elif final_score >= 40:
        final_level = "medium"
    else:
        final_level = "low"

    card["ai"]["probability"] = prob
    card["ai"]["comment"] = (res.get("comment") or "").strip() or None
    card["ai"]["risk_score"] = final_score
    card["ai"]["risk_level"] = final_level
    card["ai"]["risk_score_ai"] = risk_score_ai
    if risk_level_ai:
        card["ai"]["risk_level_ai"] = risk_level_ai

    # ONZS
    onzs_id = res.get("onzs_category_id")
    try:
        onzs_id = int(float(str(onzs_id).replace(",", ".")))
    except Exception:
        onzs_id = None
    if onzs_id and onzs_id in ONZS_CATEGORIES:
        card["onzs_category"] = onzs_id
        card["onzs_category_name"] = ONZS_CATEGORIES[onzs_id]["name"]
    elif res.get("onzs_category_name"):
        card["onzs_category_name"] = str(res.get("onzs_category_name")).strip()

    conf = _num(res.get("onzs_confidence"), None)
    if conf is not None:
        conf = max(0.0, min(1.0, conf))
    card["ai"]["onzs_confidence"] = conf

    # Recommendation + justification
    card["ai"]["recommendation"] = (res.get("recommendation") or "").strip() or None
    card["ai"]["justification"] = (res.get("justification") or "").strip() or None

    # For legacy text block compatibility (keep)
    if card.get("onzs_category_name"):
        card["ai"]["onzs_category_name"] = card.get("onzs_category_name")

def generate_card_id() -> str: 
    return str(uuid.uuid4())[:12] 
 
def save_card(card: Dict) -> str: 
    path = os.path.join(CARDS_DIR, f"{card['card_id']}.json") 
    tmp = path + ".tmp" 
    with open(tmp, "w", encoding="utf-8") as f: 
        json.dump(card, f, ensure_ascii=False, indent=2) 
    os.replace(tmp, path) 
    return path 
 
def load_card(card_id: str) -> Optional[Dict]: 
    path = os.path.join(CARDS_DIR, f"{card_id}.json") 
    if not os.path.exists(path): 
        return None 
    try: 
        with open(path, "r", encoding="utf-8") as f: 
            return json.load(f) 
    except Exception: 
        return None 
 

def build_card_text(card: Dict) -> str:
    ts = int(card.get("timestamp", now_ts()))
    dt = datetime.fromtimestamp(ts).strftime("%d.%m.%Y %H:%M")

    kw = ", ".join(card.get("keywords", [])) or "—"
    links = card.get("links") or []
    links_str = "\n".join(links) if links else "нет ссылок"

    geo = card.get("geo") or {}
    ai = card.get("ai") or {}

    # AI block (final format)
    ai_lines = []
    if ai.get("error"):
        ai_lines.append(f"🤖 Ошибка ИИ: {ai.get('error')}")
    else:
        if card.get("onzs_category_name") or ai.get("onzs_category_name"):
            ai_lines.append(f"🗂 Категория ОНзС от ИИ: {card.get('onzs_category_name') or ai.get('onzs_category_name')}")
        if ai.get("probability") is not None:
            try:
                ai_lines.append(f"📊 Вероятность самостроя: {float(ai.get('probability')):.1f}%")
            except Exception:
                ai_lines.append(f"📊 Вероятность самостроя: {ai.get('probability')}")
        if ai.get("risk_score") is not None:
            lvl = (ai.get("risk_level") or "").upper()
            ai_lines.append(f"🧠 Оценка риска: {lvl} ({int(ai.get('risk_score'))}/100)")
        if ai.get("recommendation"):
            ai_lines.append("📌 Рекомендация ИИ:")
            ai_lines.append(str(ai.get("recommendation")))
        if ai.get("justification"):
            ai_lines.append("💬 Обоснование:")
            ai_lines.append(str(ai.get("justification")))
        elif ai.get("comment"):
            ai_lines.append("💬 Комментарий ИИ:")
            ai_lines.append(str(ai.get("comment")))

    base = (
        "🔎 Обнаружено подозрительное сообщение\n"
        f"Источник: @{card.get('channel','—')}\n"
        f"Дата: {dt}\n"
        f"ID поста: {card.get('post_id','—')}\n"
    )

    if geo.get("municipality"):
        base += f"🏛 Муниципалитет: {geo.get('municipality')}\n"

    # Geo block
    geo_lines = []
    if geo.get("address"):
        geo_lines.append(f"Адрес: {geo.get('address')}")
    if geo.get("coordinates"):
        geo_lines.append(f"Координаты: {geo.get('coordinates')}")
    if geo.get("cadastral_number"):
        geo_lines.append(f"Кадастровый номер: {geo.get('cadastral_number')}")
    if geo_lines:
        base += "\n📍 Гео-информация:\n  - " + "\n  - ".join(geo_lines) + "\n"

    base += (
        f"\n🔑 Ключевые слова: {kw}\n\n"
        "📝 Текст:\n"
        f"{card.get('text','')}\n\n"
        "📎 Ссылки:\n"
        f"{links_str}\n\n"
        f"🆔 ID карточки: {card.get('card_id','—')}"
    )

    if ai_lines:
        base += "\n\n" + "\n".join(ai_lines)

    return base

def append_history(entry: Dict): 
    entry = dict(entry) 
    entry["ts"] = now_ts() 
    append_jsonl(HISTORY_CARDS, entry) 
 
 
def tg_get(method: str, params: Dict) -> Optional[Dict]: 
    if not TELEGRAM_API_URL: 
        return None 
    try: 
        r = requests.get(f"{TELEGRAM_API_URL}/{method}", params=params, timeout=HTTP_TIMEOUT) 
        return r.json() 
    except Exception as e: 
        log.error(f"Telegram GET {method} error: {e}") 
        return None 
 
def tg_post(method: str, payload: Dict) -> Optional[Dict]: 
    if not TELEGRAM_API_URL: 
        return None 
    try: 
        r = requests.post(f"{TELEGRAM_API_URL}/{method}", json=payload, timeout=HTTP_TIMEOUT) 
        return r.json() 
    except Exception as e: 
        log.error(f"Telegram POST {method} error: {e}") 
        return None 
 
def send_message(chat_id: int, text: str, reply_markup: Optional[Dict] = None) -> Optional[Dict]: 
    payload = {"chat_id": chat_id, "text": text, "disable_web_page_preview": False} 
    if reply_markup is not None: 
        payload["reply_markup"] = reply_markup 
    return tg_post("sendMessage", payload) 
 
def edit_reply_markup(chat_id: int, message_id: int, reply_markup: Optional[Dict]): 
    payload = {"chat_id": chat_id, "message_id": message_id} 
    # To remove inline keyboard for everyone, omit reply_markup field. 
    if reply_markup is not None: 
        payload["reply_markup"] = reply_markup 
    resp = tg_post("editMessageReplyMarkup", payload) 
    if resp and not resp.get("ok", True): 
        log.error(f"editMessageReplyMarkup failed: {resp}") 
    return resp 
 
 
def answer_callback(cb_id: str, text: str = "", show_alert: bool = False): 
    return tg_post("answerCallbackQuery", {"callback_query_id": cb_id, "text": text, "show_alert": show_alert}) 
 
def send_document(chat_id: int, file_path: str, filename: Optional[str] = None, caption: str = ""): 
    if not BOT_TOKEN: 
        return 
    filename = filename or os.path.basename(file_path) 
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendDocument" 
    with open(file_path, "rb") as f: 
        files = {"document": (filename, f)} 
        data = {"chat_id": chat_id, "caption": caption} 
        r = requests.post(url, data=data, files=files, timeout=HTTP_TIMEOUT) 
        if not r.ok: 
            log.error(f"sendDocument failed: {r.text}") 
 
def send_photo(chat_id: int, file_path: str, caption: str = ""): 
    if not BOT_TOKEN: 
        return 
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendPhoto" 
    with open(file_path, "rb") as f: 
        files = {"photo": (os.path.basename(file_path), f)} 
        data = {"chat_id": chat_id, "caption": caption} 
        r = requests.post(url, data=data, files=files, timeout=HTTP_TIMEOUT) 
        if not r.ok: 
            log.error(f"sendPhoto failed: {r.text}") 
 
def build_status_keyboard(card_id: str) -> Dict:
    return {
        "inline_keyboard": [
            [{"text": "На контроле у ОМС", "callback_data": f"status:{card_id}:oms_control"},
             {"text": "Направлен запрос в ОМС", "callback_data": f"status:{card_id}:oms_request"}],
            [{"text": "Дело в суде", "callback_data": f"status:{card_id}:court_case"},
             {"text": "Направлено уведомление", "callback_data": f"status:{card_id}:notification_sent"}],
        ]
    }

def build_comment_keyboard(card_id: str) -> Dict:
    return {
        "inline_keyboard": [
            [{"text": "Добавить комментарий", "callback_data": f"comment:{card_id}:add"}],
            [{"text": "Пропустить", "callback_data": f"comment:{card_id}:skip"}],
        ]
    }


ADMIN_STATE: Dict[int, str] = {}  # user_id -> pending_action

ONZS_CATEGORIES = {
    1: {"name": "Одинцовский г.о.", "stems": ["одинцов"]},
    2: {"name": "Красногорский г.о.", "stems": ["красногор"]},
    3: {"name": "Истринский г.о.", "stems": ["истринск", "истр"]},
    4: {"name": "Солнечногорский г.о.", "stems": ["солнечногор"]},
    5: {"name": "Химкинский г.о.", "stems": ["химкинск", "химк"]},
    6: {"name": "Мытищинский г.о.", "stems": ["мытищин", "мытищ"]},
    7: {"name": "Балашихинский г.о.", "stems": ["балашихин", "балаш"]},
    8: {"name": "Люберецкий г.о.", "stems": ["люберец", "любер"]},
    9: {"name": "Раменский г.о.", "stems": ["раменск"]},
    10: {"name": "Домодедовский г.о.", "stems": ["домодедов"]},
    11: {"name": "Ленинский г.о.", "stems": ["ленинск"]},
    12: {"name": "Подольский г.о.", "stems": ["подольск", "подол"]},
}

def categorize_by_location(text: str) -> Optional[int]:
    """Categorize text by location based on word stems."""
    text_lower = text.lower()
    words = set(re.findall(r'\b\w{3,}\b', text_lower))
    for cat_id, info in ONZS_CATEGORIES.items():
        for stem in info["stems"]:
            for word in words:
                if stem in word:
                    return cat_id
    return None
 



def _ru_norm(s: str) -> str:
    """Нормализация для русского текста (без агрессивной лемматизации)."""
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s

# --- Address parsing (robust) ---
_STREET_WORDS = r"(?:ул\.?|улица|пр-т\.?|проспект|просп\.?|ш\.?|шоссе|пер\.?|переулок|пл\.?|площадь|б-р\.?|бульвар|наб\.?|набережная|проезд|туп\.?|тупик)"
# Поддержка: "Октябрьская, 10" / "Октябрьская 10" / "ул Октябрьская,10" / "Октябрьская д10" / "Октябрьской 10"
_ADDR_RE_1 = re.compile(
    rf"(?:(?P<prefix>{_STREET_WORDS})\s*)?(?P<street>[А-ЯЁа-яё0-9\-\s]{3,}?)\s*[, ]\s*(?:(?:д\.?|дом)\s*)?(?P<house>\d+[А-Яа-яA-Za-z0-9\-\/]*)",
    re.IGNORECASE
)

# Более строгий вариант для "ул. X, д. 10"
_ADDR_RE_2 = re.compile(
    rf"(?P<prefix>{_STREET_WORDS})\s+(?P<street>[А-ЯЁа-яё0-9\-\s]{{3,}}?)\s*[, ]\s*(?:(?:д\.?|дом)\s*)?(?P<house>\d+[А-Яа-яA-Za-z0-9\-\/]*)",
    re.IGNORECASE
)

def _compose_address(address_raw: str, municipality_hint: Optional[str] = None) -> str:
    address_raw = _ru_norm(address_raw)
    if not address_raw:
        return ""
    # если уже содержит "Московская область" — оставляем
    if "москов" in address_raw.lower():
        return address_raw
    if municipality_hint:
        # мягкая склейка
        return f"Московская область, {municipality_hint}, {address_raw}"
    return f"Московская область, {address_raw}"

def extract_municipality_hint(text: str) -> Optional[str]:
    """
    Пытаемся вытащить округ из текста:
    - "ГО Балашиха", "г.о. Балашиха", "городской округ Балашиха"
    - просто "Балашиха" (через stems ONZS_CATEGORIES)
    """
    t = (text or "").lower()
    # явные шаблоны
    m = re.search(r"(?:г\.?\s*о\.?|го|городской\s+округ)\s*([а-яё\- ]{3,})", t, flags=re.I)
    if m:
        name = _ru_norm(m.group(1))
        if name:
            return f"г.о. {name.title()}"
    # fallback через словарь ОНзС/муниципалитетов
    cat_id = categorize_by_location(text or "")
    if cat_id:
        return ONZS_CATEGORIES.get(cat_id, {}).get("name")
    return None

# --- Object type & stage inference ---
_OBJ_TYPE_RULES = [
    ("МКД", re.compile(r"\b(жк|мкд|многоэтаж|корпус|квартир|новострой)\b", re.I)),
    ("ИЖС", re.compile(r"\b(ижс|коттедж|частн(ый|ого)\s+дом|дом\s+на\s+участке|снт|днп)\b", re.I)),
    ("Коммерция", re.compile(r"\b(тц|склад|ангар|автосервис|кафе|магазин|офис|павильон|мойка)\b", re.I)),
    ("Реконструкция", re.compile(r"\b(реконструкц|надстройк|пристройк|перепланировк)\b", re.I)),
]

_STAGE_RULES = [
    ("земляные работы", re.compile(r"\b(котлован|выемк|грунт|планировк|забор|ограждени|экскаватор|самосвал)\b", re.I)),
    ("нулевой цикл/фундамент", re.compile(r"\b(фундамент|плита|сваи|ростверк|подбетонк)\b", re.I)),
    ("надземная часть", re.compile(r"\b(этаж|монолит|каркас|колонн|перекрыти|кладк)\b", re.I)),
    ("внутренние/отделочные", re.compile(r"\b(отделк|штукатурк|перегородк|инженерн|электрик|вентиляц)\b", re.I)),
]

def infer_object_type(text: str) -> Optional[str]:
    t = text or ""
    for name, rx in _OBJ_TYPE_RULES:
        if rx.search(t):
            return name
    return None

def infer_stage(text: str) -> Optional[str]:
    t = text or ""
    for name, rx in _STAGE_RULES:
        if rx.search(t):
            return name
    return None

def detect_risk_signals(text: str) -> Dict:
    t = (text or "").lower()
    signals = {
        "mass_construction": bool(re.search(r"\b(жк|мкд|многоэтаж|комплекс|несколько\s+корпус)\b", t)),
        "danger": [],
        "complaints": bool(re.search(r"\b(жалоб|обращени|коллективн|жител[ья]|протест|митинг)\b", t)),
        "children_social": bool(re.search(r"\b(дет(и|сад)|школ|поликлиник|соцобъект)\b", t)),
        "fire_gas": bool(re.search(r"\b(пожар|газ|взрыв|задымлен)\b", t)),
        "protected_zone": bool(re.search(r"\b(охранн(ая|ой)\s+зон|лесн(ой|ого)\s+фонд|прибрежн)\b", t)),
    }
    if signals["fire_gas"]:
        signals["danger"].append("пожар/газ")
    if signals["protected_zone"]:
        signals["danger"].append("особый режим/зона")
    return signals

def rule_based_risk_score(obj_type: Optional[str], stage: Optional[str], signals: Dict, has_cadastre: bool, has_coords: bool) -> Tuple[int, str]:
    """
    Базовый скоринг (0..100) по правилам. ИИ потом может скорректировать в пределах.
    """
    score = 0
    if obj_type == "МКД":
        score += 30
    elif obj_type == "Коммерция":
        score += 15
    elif obj_type == "Реконструкция":
        score += 10
    elif obj_type == "ИЖС":
        score -= 10

    if stage == "земляные работы":
        score += 15
    elif stage == "нулевой цикл/фундамент":
        score += 10
    elif stage == "надземная часть":
        score += 20

    if signals.get("complaints"):
        score += 20
    if signals.get("children_social"):
        score += 15
    if signals.get("fire_gas"):
        score += 15
    if signals.get("protected_zone"):
        score += 25
    if signals.get("mass_construction"):
        score += 10

    if has_cadastre:
        score += 5
    if has_coords:
        score += 3

    score = max(0, min(100, int(score)))
    if score >= 70:
        level = "high"
    elif score >= 40:
        level = "medium"
    else:
        level = "low"
    return score, level


def build_admin_keyboard() -> Dict: 
    thr = get_prob_threshold() 
    return { 
        "inline_keyboard": [ 
            [{"text": f"🎯 Порог вероятности: {thr}%", "callback_data": "admin:threshold:menu"}], 
            [{"text": "📊 Статистика обучения", "callback_data": "admin:trainstats"}], 
            [{"text": "📈 График роста обучения (текст)", "callback_data": "admin:trainplot:text"}], 
            [{"text": "🖼 PNG график обучения", "callback_data": "admin:trainplot:png"}], 
            [{"text": "🗂 Журнал обучения", "callback_data": "admin:trainlog"}], 
            [{"text": "👥 Управление администраторами", "callback_data": "admin:admins:menu"}], 
            [{"text": "🧑‍⚖️ Управление модераторами", "callback_data": "admin:mods:menu"}], 
            [{"text": "🏛 Управление руководством", "callback_data": "admin:leaders:menu"}], 
            [{"text": "📄 Отчёт XLSX", "callback_data": "admin:report:xlsx"}], 
            [{"text": "🧾 Отчёт PDF", "callback_data": "admin:report:pdf"}], 
            [{"text": "📊 Дашборд KPI", "callback_data": "admin:kpi"}], 
        ] 
    } 
 
def build_threshold_keyboard() -> Dict: 
    presets = [0, 20, 40, 60, 70, 80, 90] 
    rows, row = [], [] 
    for p in presets: 
        row.append({"text": f"{p}%", "callback_data": f"admin:threshold:set:{p}"}) 
        if len(row) == 4: 
            rows.append(row); row = [] 
    if row: 
        rows.append(row) 
    rows.append([{"text": "✍️ Ввести вручную (0-100)", "callback_data": "admin:threshold:manual"}]) 
    rows.append([{"text": "⬅️ Назад", "callback_data": "admin:menu"}]) 
    return {"inline_keyboard": rows} 
 
def build_users_keyboard(kind: str) -> Dict: 
    # kind in admins/mods/leaders 
    mapping = { 
        "admins": ("администратора", "admin"), 
        "mods": ("модератора", "moderator"), 
        "leaders": ("руководство", "leadership"), 
    } 
    title, role = mapping[kind] 
    return { 
        "inline_keyboard": [ 
            [{"text": f"➕ Добавить {title}", "callback_data": f"admin:{kind}:add"}], 
            [{"text": f"➖ Удалить {title}", "callback_data": f"admin:{kind}:del"}], 
            [{"text": "📋 Показать список", "callback_data": f"admin:{kind}:list"}], 
            [{"text": "⬅️ Назад", "callback_data": "admin:menu"}], 
        ] 
    } 
 
def fetch_channel_page(url: str) -> Optional[str]: 
    headers = {"User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Safari/537.36"} 
    try: 
        r = requests.get(url, headers=headers, timeout=HTTP_TIMEOUT, allow_redirects=True) 
        if r.status_code != 200: 
            log.error(f"HTTP {r.status_code} for {url}") 
            return None 
        return r.text 
    except Exception as e: 
        log.error(f"fetch_channel_page error {url}: {e}") 
        return None 
 
def extract_posts(html: str) -> List[Dict]: 
    soup = BeautifulSoup(html, "html.parser") 
    messages = soup.find_all("div", class_="tgme_widget_message") 
    posts = [] 
    for msg in messages: 
        try: 
            msg_id = msg.get("data-post", "")  # "channel/123" 
            text_block = msg.find("div", class_="tgme_widget_message_text") 
            text = text_block.get_text(" ", strip=True) if text_block else "" 
            time_tag = msg.find("time") 
            ts = parse_tg_datetime(time_tag.get("datetime") if time_tag else "") 
            links = [] 
            for a in msg.find_all("a", href=True): 
                href = a["href"] 
                if href.startswith("http"): 
                    links.append(href) 
            posts.append({"id": msg_id, "text": text, "timestamp": ts, "links": links}) 
        except Exception as e: 
            log.error(f"extract_posts error: {e}") 
    return posts 
 
def process_channel(channel_username: str) -> List[Dict]: 
    # Forced blacklist (e.g., to exclude official channels)
    if normalize_channel(channel_username) in {normalize_channel(b) for b in BLACKLIST_CHANNELS}:
        log.info(f"[SKIP] Channel blacklisted: {channel_username}")
        return []
    url = f"https://t.me/s/{channel_username}" 
    html = fetch_channel_page(url) 
    if not html: 
        return [] 
    posts = extract_posts(html) 
    hits = [] 
    for p in posts: 
        text = normalize_text(p["text"]) 
        found = detect_keywords(text) 
        # High-signal patterns: cadastral numbers and coordinates (often in media captions) 
        if CADASTRE_RE.search(text): 
            found.append("кадастр") 
        if COORD_RE.search(text): 
            found.append("координаты") 
        # de-dup 
        found = list(dict.fromkeys([f for f in found if f])) 
        if not found: 
            continue 
        if not mark_seen(channel_username, p["id"], p["timestamp"]): 
            continue 
        hits.append({ 
            "channel": channel_username, 
            "post_id": p["id"], 
            "text": p["text"], 
            "timestamp": p["timestamp"], 
            "links": p.get("links", []), 
            "keywords": found, 
        }) 
    return hits 
 
def scan_once() -> List[Dict]: 
    all_hits: List[Dict] = [] 
    for ch in CHANNEL_LIST: 
        try: 
            hits = process_channel(ch) 
            if hits: 
                log.info(f"@{ch}: hits={len(hits)}") 
            all_hits.extend(hits) 
        except Exception as e: 
            log.error(f"scan channel @{ch} error: {e}") 
    return all_hits 
 

def extract_geo_info(text: str) -> Dict:
    """Extract geographic information from text using regex + municipality hint."""
    t = text or ""
    info: Dict = {}

    cadastre = CADASTRE_RE.search(t)
    if cadastre:
        info["cadastral_number"] = cadastre.group(0)

    coords = COORD_RE.search(t)
    if coords:
        info["coordinates"] = coords.group(0)

    municipality_hint = extract_municipality_hint(t)
    if municipality_hint:
        info["municipality_hint"] = municipality_hint

    # Address extraction: try multiple patterns; also handle "street in genitive" naturally
    addr_raw = None
    for rx in (_ADDR_RE_2, _ADDR_RE_1):
        m = rx.search(t)
        if m:
            street = _ru_norm(m.group("street"))
            house = _ru_norm(m.group("house"))
            prefix = _ru_norm(m.groupdict().get("prefix") or "")
            # Normalize prefix
            if prefix:
                prefix = prefix.replace("улица", "ул.").replace("ул", "ул.").replace("проспект", "пр-т")
                prefix = prefix.replace("пр-т.", "пр-т").replace("пр-т", "пр-т")
                prefix = prefix.replace("шоссе", "ш.").replace("ш.", "ш.")
            # If no prefix provided, keep street only
            if prefix:
                addr_raw = f"{prefix} {street}, д. {house}"
            else:
                addr_raw = f"{street}, д. {house}"
            break

    if addr_raw:
        info["address"] = _compose_address(addr_raw, municipality_hint=municipality_hint)

    return info

def enrich_geo_info(geo_info: Dict) -> Dict:
    """Enriches geo information using Yandex Geocoder API."""
    if not YANDEX_GEOCODER_API_KEY:
        return geo_info


# --- RGIS (Playwright parsing) ---
ENABLE_RGIS = str(os.getenv("ENABLE_RGIS", "1")).strip().lower() in ("1", "true", "yes", "on")
RGIS_TIMEOUT = int(os.getenv("RGIS_TIMEOUT", "35"))  # seconds
RGIS_MAX_CHARS = int(os.getenv("RGIS_MAX_CHARS", "2500"))
RGIS_HEADLESS = str(os.getenv("RGIS_HEADLESS", "1")).strip().lower() in ("1", "true", "yes", "on")

def _extract_municipality_from_rgis_text(txt: str) -> Optional[str]:
    if not txt:
        return None
    t = " ".join(str(txt).split())
    # Try common phrases
    m = re.search(r"(городской\s+округ\s+[А-Яа-яЁё\-\s]+)", t, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r"\b(г\.о\.\s*[А-Яа-яЁё\-\s]+)", t, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    # Fallback: any 'округ <Name>'
    m = re.search(r"(округ\s+[А-Яа-яЁё\-]+)", t, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return None

def rgis_fetch_planning_by_cadastre(cadastral_number: str) -> Dict:
    """
    Opens https://rgis.mosreg.ru/v3/#/?tab=planning, searches by cadastral number,
    extracts the visible result panel text. Returns dict with rgis_raw_text, municipality.
    Requires Playwright + Chromium installed in the container.
    """
    out: Dict = {"rgis_raw_text": "", "rgis_municipality": None, "rgis_ok": False, "rgis_error": None}
    cad = (cadastral_number or "").strip()
    if not cad:
        out["rgis_error"] = "empty_cadastral_number"
        return out
    if not ENABLE_RGIS:
        out["rgis_error"] = "rgis_disabled"
        return out

    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
    except Exception as e:
        out["rgis_error"] = f"playwright_import_error: {e}"
        return out

    url = "https://rgis.mosreg.ru/v3/#/?tab=planning"
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=RGIS_HEADLESS, args=["--no-sandbox"])
            ctx = browser.new_context(locale="ru-RU")
            page = ctx.new_page()
            page.set_default_timeout(RGIS_TIMEOUT * 1000)

            page.goto(url, wait_until="domcontentloaded")
            # Sometimes the app needs a moment to render
            page.wait_for_timeout(1500)

            # Try several selectors for the cadastral input
            selectors = [
                "input[placeholder*='кадастр' i]",
                "input[placeholder*='кадастров' i]",
                "input[aria-label*='кадастр' i]",
                "input[type='text']",
            ]
            inp = None
            for sel in selectors:
                loc = page.locator(sel)
                if loc.count() > 0:
                    inp = loc.first
                    try:
                        inp.click(timeout=2000)
                        break
                    except Exception:
                        continue

            if inp is None:
                raise RuntimeError("RGIS input not found")

            # Fill cadastral and search (press Enter + try click search icon)
            inp.fill("")
            inp.type(cad, delay=30)
            try:
                inp.press("Enter")
            except Exception:
                pass

            # Try click a search button/icon near input
            btn_selectors = [
                "button:has-text('Поиск')",
                "button[aria-label*='поиск' i]",
                "button:has(svg)",
            ]
            clicked = False
            for bsel in btn_selectors:
                try:
                    b = page.locator(bsel).first
                    if b.count() > 0:
                        b.click(timeout=1500)
                        clicked = True
                        break
                except Exception:
                    continue

            # Wait for any result panel to appear; collect most informative visible text.
            page.wait_for_timeout(2500)

            candidates = [
                "div:has-text('Градпроработка')",
                "div:has-text('Ограничения')",
                "div:has-text('разрешенного использования')",
                "aside",
                "section",
                "main",
            ]
            text_blocks = []
            for csel in candidates:
                try:
                    loc = page.locator(csel)
                    if loc.count() > 0:
                        # Take first few matches
                        for i in range(min(3, loc.count())):
                            t = loc.nth(i).inner_text(timeout=1500)
                            t = normalize_text(t)
                            if t and len(t) >= 40:
                                text_blocks.append(t)
                except Exception:
                    continue

            # Deduplicate and pick the longest (usually the result drawer)
            uniq = []
            seen = set()
            for t in text_blocks:
                if t in seen:
                    continue
                seen.add(t)
                uniq.append(t)
            uniq.sort(key=len, reverse=True)

            rgis_txt = (uniq[0] if uniq else "")
            if rgis_txt:
                rgis_txt = rgis_txt[:RGIS_MAX_CHARS]
                out["rgis_raw_text"] = rgis_txt
                out["rgis_ok"] = True
                out["rgis_municipality"] = _extract_municipality_from_rgis_text(rgis_txt)
            else:
                out["rgis_error"] = "rgis_no_text_found"

            try:
                ctx.close()
                browser.close()
            except Exception:
                pass

    except Exception as e:
        out["rgis_error"] = str(e)

    return out

    if "address" in geo_info and "coordinates" not in geo_info:
        try:
            url = f"https://geocode-maps.yandex.ru/1.x/?apikey={YANDEX_GEOCODER_API_KEY}&format=json&geocode={geo_info['address']}"
            r = requests.get(url, timeout=HTTP_TIMEOUT)
            data = r.json()
            pos = data["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]["Point"]["pos"]
            geo_info["coordinates"] = ", ".join(pos.split(" ")[::-1])
        except Exception as e:
            log.error(f"Yandex Geocoder (forward) error: {e}")

    elif "coordinates" in geo_info and "address" not in geo_info:
        try:
            url = f"https://geocode-maps.yandex.ru/1.x/?apikey={YANDEX_GEOCODER_API_KEY}&format=json&geocode={geo_info['coordinates'].replace(' ', '')}"
            r = requests.get(url, timeout=HTTP_TIMEOUT)
            data = r.json()
            address = data["response"]["GeoObjectCollection"]["featureMember"][0]["GeoObject"]["metaDataProperty"]["GeocoderMetaData"]["text"]
            geo_info["address"] = address
        except Exception as e:
            log.error(f"Yandex Geocoder (reverse) error: {e}")
            
    return geo_info


# --- Cadastre lookup fallback (no Playwright) ---
# Uses public PKK (Rosreestr) endpoints when available. If the endpoint changes,
# the scraper will just skip enrichment without failing the whole card.
def lookup_cadastre_pkk(cn: str) -> Dict:
    """
    Try to resolve cadastral number to an address and basic attributes via PKK (Rosreestr).
    Returns dict with keys: address, raw (optional).
    """
    cn = (cn or "").strip()
    if not cn:
        return {}
    try:
        # search object by text
        s_url = "https://pkk.rosreestr.ru/api/features/1"
        r = requests.get(s_url, params={"text": cn}, timeout=HTTP_TIMEOUT)
        if not r.ok:
            return {}
        js = r.json()
        feats = (js or {}).get("features") or []
        if not feats:
            return {}
        fid = feats[0].get("attrs", {}).get("id") or feats[0].get("id")
        if not fid:
            return {}
        d_url = f"https://pkk.rosreestr.ru/api/features/1/{fid}"
        r2 = requests.get(d_url, timeout=HTTP_TIMEOUT)
        if not r2.ok:
            return {}
        js2 = r2.json() or {}
        attrs = (js2.get("feature") or {}).get("attrs") or {}
        # different schemas exist; best-effort
        addr = attrs.get("address") or attrs.get("readable_address") or attrs.get("location") or ""
        out = {}
        if addr:
            out["address"] = str(addr).strip()
        out["pkk_raw"] = attrs
        return out
    except Exception as e:
        log.error(f"PKK cadastre lookup error: {e}")
        return {}

def enrich_geo_info_fallback(geo_info: Dict) -> Dict:
    """
    Fallback enrichment without Playwright:
    - If cadastre exists and address is missing -> try PKK to get address.
    - If we got address -> forward geocode to coords.
    """
    if not isinstance(geo_info, dict):
        return {}
    cn = geo_info.get("cadastral_number")
    if cn and not geo_info.get("address"):
        pkk = lookup_cadastre_pkk(cn)
        if pkk.get("address"):
            geo_info["address"] = pkk["address"]
        geo_info.setdefault("sources", {})
        geo_info["sources"]["pkk"] = True if pkk else False

    # Reuse existing Yandex geocoder enrichment for coords/address
    try:
        geo_info = enrich_geo_info(geo_info)
    except Exception:
        pass
    return geo_info

def generate_card(hit: Dict) -> Dict:
    """
    Builds a unified "samostroy profile" card (canonical JSON).
    Keeps legacy top-level fields for backward compatibility with the existing bot logic.
    """
    cid = generate_card_id()
    text_raw = hit.get("text", "") or ""
    keywords = hit.get("keywords") or []

    # geo extraction + enrichment
    geo_info = extract_geo_info(text_raw)
    geo_info = enrich_geo_info(geo_info)

    municipality = geo_info.get("municipality_hint") or None
    cad = geo_info.get("cadastral_number")
    coords = geo_info.get("coordinates")
    address = geo_info.get("address")

    # object inference
    obj_type = infer_object_type(text_raw)
    stage = infer_stage(text_raw)
    signals = detect_risk_signals(text_raw)

    # rule-based risk baseline
    base_score, base_level = rule_based_risk_score(
        obj_type=obj_type,
        stage=stage,
        signals=signals,
        has_cadastre=bool(cad),
        has_coords=bool(coords),
    )

    # ONZS via location heuristic (soft)
    category_id = categorize_by_location(text_raw)
    onzs_name = ONZS_CATEGORIES.get(category_id, {}).get("name") if category_id else None
    if municipality and not onzs_name:
        # municipality_hint often already matches ONZS dict
        for k, v in ONZS_CATEGORIES.items():
            if (v.get("name") or "").lower() == (municipality or "").lower():
                category_id = k
                onzs_name = v.get("name")
                break

    # Canonical profile
    card: Dict = {
        "card_id": cid,

        # legacy
        "channel": hit.get("channel"),
        "post_id": hit.get("post_id"),
        "timestamp": hit.get("timestamp"),
        "text": text_raw,
        "keywords": keywords,
        "links": hit.get("links", []),

        # canonical sections
        "source": {
            "channel": hit.get("channel"),
            "post_id": hit.get("post_id"),
            "published_ts": hit.get("timestamp"),
            "credibility": "medium",
        },
        "text_profile": {
            "raw": text_raw,
            "keywords": keywords,
            "signals": signals,
        },
        "geo": {
            "address": address,
            "municipality": municipality,
            "coordinates": coords,
            "cadastral_number": cad,
            "geo_source": geo_info.get("geo_source") or ("yandex_geocoder" if (address or coords) else None),
        },
        "object": {
            "type": obj_type,
            "stage": stage,
            "floors_estimated": None,
            "area_estimated": None,
        },
        "legal": {
            "assumed_violation": [],
            "special_zone": [],
        },
        "onzs": {
            "category_id": category_id,
            "category_name": onzs_name,
            "source": "geo-heuristic",
        },
        "ai": {
            "probability": None,
            "comment": None,
            "risk_level_rule": base_level,
            "risk_score_rule": base_score,
            "risk_level": base_level,
            "risk_score": base_score,
            "recommendation": None,
            "justification": None,
            "model": "rules",
            "confidence": None,
        },
        "workflow": {
            "status": "in_review",
            "priority": "high" if base_level == "high" else ("medium" if base_level == "medium" else "low"),
            "history": [],
        },

        # keep old fields used elsewhere
        "geo_info": geo_info,
        "status": "new",
        "history": [],
    }

    # keep top-level onzs fields (legacy)
    if category_id:
        card["onzs_category"] = category_id
        card["onzs_category_name"] = ONZS_CATEGORIES[category_id]["name"]

    # AI enrichment (may update onzs + risk + probability)
    try:
        enrich_card_with_yagpt(card)
        # sync canonical ONZS object from top-level if AI set them
        if card.get("onzs_category"):
            card["onzs"]["category_id"] = card.get("onzs_category")
            card["onzs"]["category_name"] = card.get("onzs_category_name")
            card["onzs"]["source"] = "ai+geo"
        # sync ai model name
        if card.get("ai") is not None:
            card["ai"]["model"] = "yandexgpt"
    except Exception as e:
        log.error(f"enrich_card_with_yagpt error: {e}")

    save_card(card)
    return card

def send_card_to_group(card: Dict) -> Optional[int]: 
    thr = get_prob_threshold() 
    prob = None 
    try: 
        prob = float((card.get("ai") or {}).get("probability")) 
    except Exception: 
        prob = None 
 
    if prob is not None and prob < thr: 
        card["status"] = "filtered" 
        card.setdefault("history", []).append({"event": "filtered", "threshold": thr, "ts": now_ts()}) 
        save_card(card) 
        append_history({"event": "filtered", "card_id": card["card_id"], "threshold": thr, "prob": prob}) 
        return None 
 
    res = send_message(TARGET_CHAT_ID, build_card_text(card), reply_markup=build_card_keyboard(card["card_id"])) 
    if not res or not res.get("ok"): 
        log.error(f"sendMessage failed: {res}") 
        return None 
 
    msg = res["result"] 
    card.setdefault("tg", {}) 
    card["tg"]["chat_id"] = msg["chat"]["id"] 
    card["tg"]["message_id"] = msg["message_id"] 
    card["status"] = "sent" 
    card.setdefault("history", []).append({"event": "sent", "ts": now_ts(), "chat_id": card["tg"]["chat_id"], "message_id": card["tg"]["message_id"]}) 
    save_card(card) 
    append_history({"event": "sent", "card_id": card["card_id"], "chat_id": card["tg"]["chat_id"], "message_id": card["tg"]["message_id"]}) 
    return msg["message_id"] 
 
def apply_card_action(card_id: str, action: str, from_user: int) -> Tuple[str, bool]:
    """
    Returns (message, decided_now).
    decided_now=True only for the first admin that made the decision.
    """
    existing = decision_exists(card_id)
    if existing:
        dec, by, ts = existing
        dt = datetime.fromtimestamp(ts).strftime("%d.%m.%Y %H:%M")
        return (f"Уже обработано: {dec} (админ {by}, {dt})", False)

    if action not in ("work", "wrong"):
        return ("Неизвестное действие.", False)

    card = load_card(card_id)
    if not card:
        return ("Карточка не найдена.", False)

    wrote = set_decision(card_id, action, from_user)
    if not wrote:
        return ("Уже обработано другим администратором.", False)

    old_status = card.get("status", "new")
    if action == "work":
        # The 'work' action now triggers the next step in the flow, handled in handle_callback_query
        return ("Выберите статус:", True)
    elif action == "wrong":
        new_status, label, msg = "wrong", "wrong", "Статус: НЕВЕРНО ❌"
    card["status"] = new_status
    card.setdefault("history", []).append({"event": f"set_{new_status}", "from_user": int(from_user), "ts": now_ts()})
    save_card(card)

    append_history({"event": "status_change", "card_id": card_id, "from_user": int(from_user), "old_status": old_status, "new_status": new_status})
    log_training_event(card_id, label, card.get("text", ""), card.get("channel", ""), admin_id=int(from_user))
    return (msg, True)


def _fetch_train_daily_last(days: int = 30): 
    conn = db() 
    rows = conn.execute("SELECT day, total, work, wrong, attach FROM train_daily ORDER BY day DESC LIMIT ?;", (int(days),)).fetchall() 
    conn.close() 
    return list(reversed(rows)) 
 
def build_kpi_text() -> str: 
    rows = _fetch_train_daily_last(30) 
    if not rows: 
        return "📊 KPI: данных обучения пока нет." 
    total = sum(int(r[1]) for r in rows) 
    work = sum(int(r[2]) for r in rows) 
    wrong = sum(int(r[3]) for r in rows)    acc = (work / total * 100.0) if total > 0 else 0.0 
    last_day = rows[-1][0] 
    return ( 
        "📊 KPI (самострой-контроль)\n" 
        f"Период: последние {len(rows)} дн. (до {last_day})\n\n" 
        f"Всего решений: {total}\n" 
        f"В работу: {work}\n" 
        f"Неверно: {wrong}\n" 
                f"Доля полезных (в работу+привязать): {acc:.1f}%\n" 
    ) 
 

def _iter_cards_in_period(ts_from: int, ts_to: int) -> List[Dict]:
    """Loads card JSONs from CARDS_DIR and returns those within [ts_from, ts_to]."""
    out: List[Dict] = []
    try:
        for fn in os.listdir(CARDS_DIR):
            if not fn.endswith(".json"):
                continue
            fp = os.path.join(CARDS_DIR, fn)
            try:
                with open(fp, "r", encoding="utf-8") as f:
                    c = json.load(f)
                ts = int(c.get("timestamp") or c.get("source", {}).get("published_ts") or 0)
                if ts_from <= ts <= ts_to:
                    out.append(c)
            except Exception:
                continue
    except Exception:
        pass
    return out

def _period_bounds(days: int) -> Tuple[int, int, str]:
    """UTC bounds for last N days (inclusive)."""
    now = int(time.time())
    ts_to = now
    ts_from = now - int(days) * 86400
    label = f"последние {days} дн."
    return ts_from, ts_to, label

def compute_analytics(cards: List[Dict]) -> Dict:
    """
    Computes analytics for leadership:
    - totals
    - by object type, stage
    - by risk_level
    - by onzs_category_name
    """
    def _get(obj, *path, default=None):
        cur = obj
        for p in path:
            if not isinstance(cur, dict):
                return default
            cur = cur.get(p)
        return cur if cur is not None else default

    total = len(cards)

    by_type = {}
    by_stage = {}
    by_risk = {"high": 0, "medium": 0, "low": 0, "unknown": 0}
    by_onzs = {}

    for c in cards:
        obj_type = _get(c, "object", "type") or _get(c, "object", "type", default=None) or _get(c, "object", "type") or None
        stage = _get(c, "object", "stage") or None
        risk = (_get(c, "ai", "risk_level") or "").lower() or "unknown"
        onzs = c.get("onzs_category_name") or _get(c, "onzs", "category_name") or "—"

        by_type[obj_type or "—"] = by_type.get(obj_type or "—", 0) + 1
        by_stage[stage or "—"] = by_stage.get(stage or "—", 0) + 1
        if risk not in by_risk:
            by_risk["unknown"] += 1
        else:
            by_risk[risk] += 1
        by_onzs[onzs] = by_onzs.get(onzs, 0) + 1

    # Sort dicts
    by_type = dict(sorted(by_type.items(), key=lambda x: x[1], reverse=True))
    by_stage = dict(sorted(by_stage.items(), key=lambda x: x[1], reverse=True))
    by_onzs = dict(sorted(by_onzs.items(), key=lambda x: x[1], reverse=True))

    return {
        "total": total,
        "by_type": by_type,
        "by_stage": by_stage,
        "by_risk": by_risk,
        "by_onzs": by_onzs,
    }

def build_leadership_summary_text(days: int = 14) -> str:
    ts_from, ts_to, label = _period_bounds(days)
    cards = _iter_cards_in_period(ts_from, ts_to)
    a = compute_analytics(cards)

    # Pretty
    def fmt_top(d: Dict, n: int = 5) -> str:
        items = list(d.items())[:n]
        if not items:
            return "—"
        return "\n".join([f"• {k}: {v}" for k, v in items])

    return (
        "🧾 Сводка по самострою (для руководства)\n"
        f"Период: {label}\n\n"
        f"Всего выявлено карточек: {a['total']}\n\n"
        "🏗 По типу объекта:\n" + fmt_top(a["by_type"], 8) + "\n\n"
        "🧱 По стадии:\n" + fmt_top(a["by_stage"], 8) + "\n\n"
        "🚦 По уровню риска:\n"
        f"• HIGH: {a['by_risk'].get('high',0)}\n"
        f"• MEDIUM: {a['by_risk'].get('medium',0)}\n"
        f"• LOW: {a['by_risk'].get('low',0)}\n"
        f"• UNKNOWN: {a['by_risk'].get('unknown',0)}\n\n"
        "🏛 Топ ОНзС:\n" + fmt_top(a["by_onzs"], 10)
    )



def build_report_xlsx() -> str:
    """
    Leadership-ready XLSX report:
    - Summary
    - KPI (training)
    - Analytics (samostroy cards) for last REPORT_DAYS
    - TrainingDaily + ChannelBias
    - WorkReport_ONZS (status history)
    """
    report_days = int(os.getenv("REPORT_DAYS", "14") or 14)
    out_path = os.path.join(REPORTS_DIR, f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb = Workbook()

    # === Summary sheet (leadership) ===
    ws0 = wb.active
    ws0.title = "Summary"
    ws0.append(["Раздел", "Значение"])
    summary_text = build_leadership_summary_text(days=report_days)
    for ln in summary_text.splitlines():
        if not ln.strip():
            continue
        if ":" in ln and not ln.strip().startswith("http"):
            k, v = ln.split(":", 1)
            ws0.append([k.strip("• ").strip(), v.strip()])
        else:
            ws0.append([ln.strip(), ""])

    # === KPI sheet (training) ===
    ws = wb.create_sheet("KPI_Training")
    ws.append(["Показатель", "Значение"])
    for line in build_kpi_text().splitlines()[1:]:
        if ":" in line:
            k, v = line.split(":", 1)
            ws.append([k.strip(), v.strip()])

    # === Analytics sheets (cards) ===
    ts_from, ts_to, label = _period_bounds(report_days)
    cards = _iter_cards_in_period(ts_from, ts_to)
    a = compute_analytics(cards)

    wsA = wb.create_sheet("Analytics_Cards")
    wsA.append(["Период", label])
    wsA.append(["Всего карточек", a["total"]])
    wsA.append([])
    wsA.append(["Срез", "Значение", "Кол-во"])

    for k, v in a["by_type"].items():
        wsA.append(["Тип объекта", k, v])
    wsA.append([])
    for k, v in a["by_stage"].items():
        wsA.append(["Стадия", k, v])
    wsA.append([])
    for k, v in a["by_risk"].items():
        wsA.append(["Риск", k, v])
    wsA.append([])
    for k, v in a["by_onzs"].items():
        wsA.append(["ОНзС", k, v])

    # === TrainingDaily ===
    ws2 = wb.create_sheet("TrainingDaily")
    ws2.append(["day", "total", "work", "wrong", "attach"])
    for r in _fetch_train_daily_last(90):
        ws2.append(list(r))

    # === ChannelBias ===
    ws3 = wb.create_sheet("ChannelBias")
    ws3.append(["channel", "bias_points"])
    w = _get_model_param("weights", {"channels": {}})
    for ch, b in sorted((w.get("channels") or {}).items(), key=lambda x: x[0]):
        ws3.append([ch, b])

    # === WorkReport_ONZS ===
    ws4 = wb.create_sheet("WorkReport_ONZS")
    headers = ["Card ID", "Channel", "Post ID", "Timestamp", "Category", "Final Status", "Comment", "Last Updated By", "Last Updated Ts"]
    ws4.append(headers)

    conn = db()
    rows = conn.execute("""
        SELECT
            cs.card_id,
            cs.onzs_category,
            cs.status,
            cs.comment,
            cs.last_updated_by,
            cs.last_updated_ts
        FROM card_status cs
        ORDER BY cs.last_updated_ts DESC
    """).fetchall()
    conn.close()

    for row in rows:
        card_id, onzs_category, status, comment, updated_by, updated_ts = row
        card_data = load_card(card_id)
        if card_data:
            channel = card_data.get("channel", "")
            post_id = card_data.get("post_id", "")
            timestamp = datetime.fromtimestamp(int(card_data.get("timestamp", 0) or 0)).strftime("%Y-%m-%d %H:%M:%S")
            category_name = ONZS_CATEGORIES.get(onzs_category, {}).get("name", "N/A") if onzs_category else (card_data.get("onzs_category_name") or "N/A")
            updated_ts_str = datetime.fromtimestamp(int(updated_ts or 0)).strftime("%Y-%m-%d %H:%M:%S")
            ws4.append([card_id, channel, post_id, timestamp, category_name, status, comment, updated_by, updated_ts_str])

    # Formatting: basic widths
    for wsx in [ws0, ws, wsA, ws2, ws3, ws4]:
        for col in range(1, wsx.max_column + 1):
            wsx.column_dimensions[get_column_letter(col)].width = 32

    wb.save(out_path)
    return out_path


def build_report_pdf() -> str:
    report_days = int(os.getenv("REPORT_DAYS", "14") or 14)
    out_path = os.path.join(REPORTS_DIR, f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf")
    c = canvas.Canvas(out_path, pagesize=A4)
    width, height = A4

    text = c.beginText(40, height - 60)
    text.setFont("Helvetica", 11)

    # Leadership summary
    for line in build_leadership_summary_text(days=report_days).splitlines():
        text.textLine(line)

    text.textLine("")
    text.textLine("—" * 60)
    text.textLine("")

    # Training KPI
    for line in build_kpi_text().splitlines():
        text.textLine(line)

    c.drawText(text)
    c.showPage()
    c.save()
    return out_path

def build_trainplot_png(days: int = 60) -> str: 
    out_path = os.path.join(REPORTS_DIR, f"trainplot_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.png") 
    rows = _fetch_train_daily_last(days) 
    plt.figure(figsize=(10, 4)) 
    if not rows: 
        plt.title("Training (no data)") 
        plt.savefig(out_path, dpi=150, bbox_inches="tight") 
        plt.close() 
        return out_path 
 
    days_list = [r[0] for r in rows] 
    total = [int(r[1]) for r in rows] 
    work = [int(r[2]) for r in rows] 
    wrong = [int(r[3]) for r in rows]    plt.plot(days_list, total, label="total") 
    plt.plot(days_list, work, label="work") 
    plt.plot(days_list, wrong, label="wrong") 
plt.xticks(rotation=45, ha="right") 
    plt.legend() 
    plt.tight_layout() 
    plt.savefig(out_path, dpi=150, bbox_inches="tight") 
    plt.close() 
    return out_path 
 
def get_all_report_recipients() -> List[int]: 
    ids = set() 
    for role in ("leadership", "admin", "moderator"): 
        for uid in list_users_by_role(role): 
            ids.add(int(uid)) 
    return sorted(ids) 
 
def daily_reports_worker(): 
    # Daily at 09:00 Moscow 
    try: 
        from zoneinfo import ZoneInfo 
        tz = ZoneInfo("Europe/Moscow") 
    except Exception: 
        tz = None 
 
    while True: 
        now = datetime.now(tz) if tz else datetime.now() 
        target = now.replace(hour=9, minute=0, second=0, microsecond=0) 
        if target <= now: 
            target = target + timedelta(days=1) 
        time.sleep(max(5, int((target - now).total_seconds()))) 
 
        try: 
            kpi = build_kpi_text() 
            xlsx = build_report_xlsx() 
            pdf = build_report_pdf() 
            png = build_trainplot_png() 
 
            for uid in get_all_report_recipients(): 
                send_message(uid, kpi) 
                send_document(uid, xlsx, caption="📄 Ежедневный отчёт (XLSX)") 
                send_document(uid, pdf, caption="🧾 Ежедневный отчёт (PDF)") 
                send_photo(uid, png, caption="📈 График обучения") 
        except Exception as e: 
            log.exception(f"daily_reports_worker error: {e}") 
 
UPDATE_OFFSET = get_update_offset()
LAST_CONFLICT_ALERT_TS = 0
 
def handle_callback_query(upd: Dict):
    cb = upd.get("callback_query") or {}
    cb_id = cb.get("id")
    from_user = int((cb.get("from") or {}).get("id", 0))
    data = (cb.get("data") or "").strip()
    msg_obj = cb.get("message") or {}
    chat_id = (msg_obj.get("chat") or {}).get("id")
    message_id = msg_obj.get("message_id")

    role = get_role(from_user)

    if data.startswith("card:"):
        if not is_admin(from_user):
            answer_callback(cb_id, "Только администраторы могут менять статус.", show_alert=True)
            return
        _, card_id, action = data.split(":", 2)

        # Backward compatibility: old messages may still have "attach" button
        if action == "attach":
            if chat_id and message_id:
                edit_reply_markup(chat_id, message_id, reply_markup=None)
            answer_callback(cb_id, "Кнопка «Привязать» отключена. Используйте «В работу» или «Неверно».", show_alert=True)
            return
        card = load_card(card_id)

        if not card:
            answer_callback(cb_id, "Карточка не найдена.", show_alert=True)
            return

        result, decided_now = apply_card_action(card_id, action, from_user)

        if decided_now and action == "work":
            edit_reply_markup(chat_id, message_id, reply_markup=build_status_keyboard(card_id))
            answer_callback(cb_id, "Выберите статус", show_alert=False)
        else:
            if chat_id and message_id:
                edit_reply_markup(chat_id, message_id, reply_markup=None)
            answer_callback(cb_id, result, show_alert=False)
        return

    if data.startswith("status:"):
        if not is_admin(from_user):
            answer_callback(cb_id, "Только администраторы могут менять статус.", show_alert=True)
            return
        _, card_id, status = data.split(":", 2)
        card = load_card(card_id)
        if not card:
            answer_callback(cb_id, "Карточка не найдена.", show_alert=True)
            return

        # Persist final status (this is the "end" of the workflow for label=work)
        conn = db()
        conn.execute(
            "INSERT OR REPLACE INTO card_status (card_id, onzs_category, status, last_updated_ts, last_updated_by) VALUES (?, ?, ?, ?, ?)",
            (card_id, int(card.get("onzs_category") or 0) or None, status, now_ts(), from_user)
        )
        conn.close()

        # Update card file and write training event at FINAL status selection
        try:
            old_status = card.get("status", "new")
            card["status"] = "work"
            card.setdefault("history", []).append({"event": "final_status_selected", "final_status": status, "from_user": int(from_user), "ts": now_ts()})
            save_card(card)
            append_history({"event": "final_status_selected", "card_id": card_id, "from_user": int(from_user), "old_status": old_status, "final_status": status})
            log_training_event(card_id, "work", card.get("text", ""), card.get("channel", ""), admin_id=int(from_user))
        except Exception as e:
            log.error(f"final status handling error: {e}")

        edit_reply_markup(chat_id, message_id, reply_markup=build_comment_keyboard(card_id))
        answer_callback(cb_id, f"Статус '{status}' установлен.", show_alert=False)
        return

    if data.startswith("comment:"):
        if not is_admin(from_user):
            answer_callback(cb_id, "Только администраторы могут менять статус.", show_alert=True)
            return
        _, card_id, action = data.split(":", 2)

        # Backward compatibility: old messages may still have "attach" button
        if action == "attach":
            if chat_id and message_id:
                edit_reply_markup(chat_id, message_id, reply_markup=None)
            answer_callback(cb_id, "Кнопка «Привязать» отключена. Используйте «В работу» или «Неверно».", show_alert=True)
            return
        if action == "add":
            ADMIN_STATE[from_user] = f"await_comment:{card_id}"
            send_message(chat_id, "Пожалуйста, введите комментарий для карточки.")
            answer_callback(cb_id, "Ожидаю ваш комментарий.", show_alert=False)
        else: # skip
            edit_reply_markup(chat_id, message_id, reply_markup=None)
            send_message(chat_id, f"Карточка {card_id} обработана.")
            answer_callback(cb_id, "Обработка завершена.", show_alert=False)
        return

    if data.startswith("admin:"):
        if not is_admin(from_user): 
            answer_callback(cb_id, "❌ Нет доступа.", show_alert=True) 
            return 
 
        parts = data.split(":") 
 
        if data == "admin:menu": 
            send_message(chat_id, "🛠 Админ-панель:", reply_markup=build_admin_keyboard()) 
            answer_callback(cb_id, "Ок"); return 
 
        # Threshold 
        if data == "admin:threshold:menu": 
            send_message(chat_id, "🎯 Настройка порога вероятности (0–100):", reply_markup=build_threshold_keyboard()) 
            answer_callback(cb_id, "Ок"); return 
 
        if len(parts) == 4 and parts[1] == "threshold" and parts[2] == "set": 
            try: v = int(parts[3]) 
            except Exception: v = DEFAULT_THRESHOLD 
            set_prob_threshold(v) 
            send_message(chat_id, f"✅ Порог установлен: {get_prob_threshold()}%", reply_markup=build_admin_keyboard()) 
            answer_callback(cb_id, "Сохранено"); return 
 
        if data == "admin:threshold:manual": 
            ADMIN_STATE[from_user] = "await_threshold" 
            send_message(chat_id, "Введите порог числом 0–100 (сообщением).") 
            answer_callback(cb_id, "Ожидаю ввод"); return 
 
        # Users management 
        if data == "admin:admins:menu": 
            send_message(chat_id, "👥 Управление администраторами:", reply_markup=build_users_keyboard("admins")) 
            answer_callback(cb_id, "Ок"); return 
        if data == "admin:mods:menu": 
            send_message(chat_id, "🧑‍⚖️ Управление модераторами:", reply_markup=build_users_keyboard("mods")) 
            answer_callback(cb_id, "Ок"); return 
        if data == "admin:leaders:menu": 
            send_message(chat_id, "🏛 Управление руководством:", reply_markup=build_users_keyboard("leaders")) 
            answer_callback(cb_id, "Ок"); return 
 
        # list/add/del handlers 
        if len(parts) == 3 and parts[2] == "list" and parts[1] in ("admins","mods","leaders"): 
            role_map = {"admins":"admin","mods":"moderator","leaders":"leadership"} 
            role_key = role_map[parts[1]] 
            ids = list_users_by_role(role_key) 
            txt = "\n".join(str(i) for i in ids) if ids else "Список пуст." 
            send_message(chat_id, f"Список ({role_key}):\n{txt}") 
            answer_callback(cb_id, "Ок"); return 
 
        if len(parts) == 3 and parts[2] in ("add","del") and parts[1] in ("admins","mods","leaders"): 
            op = parts[2] 
            ADMIN_STATE[from_user] = f"await_{op}_{parts[1]}" 
            send_message(chat_id, "Отправьте Telegram ID (числом) следующим сообщением.") 
            answer_callback(cb_id, "Ожидаю ID"); return 
 
        # Reports & KPI 
        if data == "admin:report:xlsx": 
            p = build_report_xlsx() 
            send_document(chat_id, p, caption="📄 Отчёт (XLSX)") 
            answer_callback(cb_id, "Готово"); return 
 
        if data == "admin:report:pdf": 
            p = build_report_pdf() 
            send_document(chat_id, p, caption="🧾 Отчёт (PDF)") 
            answer_callback(cb_id, "Готово"); return 
 
        if data == "admin:kpi": 
            send_message(chat_id, build_kpi_text()) 
            answer_callback(cb_id, "Ок"); return 
 
        # Training info 
        if data == "admin:trainstats": 
            st = compute_training_stats() 
            last = st.get("last_ts") 
            last_s = st.get("last_str") or (datetime.fromtimestamp(last).strftime("%d.%m.%Y %H:%M") if last else "—") 
            send_message( 
                chat_id, 
                "📊 Статистика обучения (агрегация по всем админам):\n\n" 
                f"• Всего событий: {st['total']}\n" 
                f"   ├─ В работу: {st['work']}\n" 
                f"   └─ Неверно: {st['wrong']}\n\n" 
                f"• Прогресс к цели ({st['target']}): {st['progress']}%\n" 
                f"• Условная уверенность: {st['confidence']}%\n" 
                f"• Последнее событие: {last_s}\n" 
            ) 
            answer_callback(cb_id, "Ок"); return 
 
        if data == "admin:trainplot:text": 
            send_message(chat_id, training_plot_text(days=14)) 
            answer_callback(cb_id, "Ок"); return 
 
        if data == "admin:trainplot:png": 
            p = build_trainplot_png() 
            send_photo(chat_id, p, caption="📈 График обучения (PNG)") 
            answer_callback(cb_id, "Ок"); return 
 
        if data == "admin:trainlog": 
            events = tail_training_log(limit=MAX_TRAIN_LOG) 
            if not events: 
                send_message(chat_id, "🗂 Журнал обучения пуст.") 
                answer_callback(cb_id, "Ок"); return 
            lines = ["🗂 Последние события обучения:"] 
            for e in events[-MAX_TRAIN_LOG:]: 
                ts = e.get("timestamp") 
                dt = datetime.fromtimestamp(int(ts)).strftime("%d.%m %H:%M") if isinstance(ts, int) else "—" 
                lbl = e.get("label", "—") 
                adm = e.get("admin_id", "—") 
                cid = e.get("card_id", "—") 
                ch = e.get("channel", "—") 
                lines.append(f"• {dt} | {lbl} | @{ch} | admin={adm} | card={cid}") 
            send_message(chat_id, "\n".join(lines)) 
            answer_callback(cb_id, "Ок"); return 
 
        answer_callback(cb_id, "Неизвестное действие.", show_alert=False) 
        return 
 
    answer_callback(cb_id, "") 
 
def handle_message(upd: Dict): 
    # message sources: private/group message, edited_message, channel_post, edited_channel_post 
    msg = (upd.get("message") or upd.get("edited_message") or upd.get("channel_post") or upd.get("edited_channel_post") or {}) 
    chat_id = (msg.get("chat") or {}).get("id") 
    from_user = int((msg.get("from") or {}).get("id", 0)) 
    # commands/text can be in text or caption (media posts) 
    text = ((msg.get("text") or msg.get("caption") or "")).strip() 
 
    # stateful admin inputs
    if is_admin(from_user) and from_user in ADMIN_STATE and not text.startswith("/"):
        st = ADMIN_STATE.pop(from_user, "")

        if st == "await_threshold":
            m = re.findall(r"-?\d+", text)
            if not m:
                send_message(chat_id, "❌ Не распознал число. Введите 0–100.")
                ADMIN_STATE[from_user] = "await_threshold"
                return
            set_prob_threshold(int(m[0]))
            send_message(chat_id, f"✅ Порог установлен: {get_prob_threshold()}%")
            return

        if st.startswith("await_comment:"):
            try:
                _, card_id = st.split(":", 1)
                conn = db()
                conn.execute(
                    "UPDATE card_status SET comment = ?, last_updated_ts = ?, last_updated_by = ? WHERE card_id = ?",
                    (text, now_ts(), from_user, card_id)
                )
                conn.close()
                send_message(chat_id, f"✅ Комментарий для карточки {card_id} добавлен.")
                # Attempt to remove the keyboard from the original message if possible, though message_id isn't stored.
                # This part is best-effort. The main confirmation is the message above.
            except Exception as e:
                log.error(f"Error adding comment: {e}")
                send_message(chat_id, "❌ Ошибка при добавлении комментария.")
            return

        # user role operations
        m = re.findall(r"\d+", text)
        if not m:
            send_message(chat_id, "❌ Не распознал ID. Отправьте число.")
            ADMIN_STATE[from_user] = st
            return
        uid = int(m[0])

        if st == "await_add_admins":
            add_admin(uid); send_message(chat_id, f"✅ Администратор добавлен: {uid}", reply_markup=build_admin_keyboard()); return
        if st == "await_del_admins":
            if uid == from_user:
                send_message(chat_id, "❌ Нельзя удалить самого себя через меню. Используйте другого админа."); return
            remove_admin(uid); send_message(chat_id, f"🗑 Администратор удалён: {uid}", reply_markup=build_admin_keyboard()); return

        if st == "await_add_mods":
            add_moderator(uid); send_message(chat_id, f"✅ Модератор добавлен: {uid}", reply_markup=build_admin_keyboard()); return
        if st == "await_del_mods":
            remove_moderator(uid); send_message(chat_id, f"🗑 Модератор удалён: {uid}", reply_markup=build_admin_keyboard()); return

        if st == "await_add_leaders":
            add_leadership(uid); send_message(chat_id, f"✅ Добавлено в руководство: {uid}", reply_markup=build_admin_keyboard()); return
        if st == "await_del_leaders":
            remove_leadership(uid); send_message(chat_id, f"🗑 Удалено из руководства: {uid}", reply_markup=build_admin_keyboard()); return

        # unknown state
        send_message(chat_id, "⚠️ Неизвестная операция. /admin")
        return

    if not text.startswith("/"):
        return

    cmd = text.split()[0].split("@")[0]
    log.info(f"[CMD] {cmd} from_user={from_user} chat_id={chat_id}")

    if cmd == "/get_work_report":
        if not (is_admin(from_user) or is_leadership(from_user)):
            send_message(chat_id, "❌ Нет доступа.")
            return
        try:
            report_path = build_report_xlsx()
            send_document(chat_id, report_path, caption="📄 Отчет о работе по ОНзС")
        except Exception as e:
            log.error(f"Failed to generate work report: {e}")
            send_message(chat_id, "❌ Не удалось создать отчет.")
        return

    if cmd == "/admin":
        if not is_admin(from_user): 
            send_message(chat_id, "❌ Команда /admin доступна только администраторам.") 
            return 
        send_message(chat_id, "🛠 Админ-панель:", reply_markup=build_admin_keyboard()) 
        return 
 
    if cmd == "/dashboard": 
        if not (is_admin(from_user) or is_leadership(from_user)): 
            send_message(chat_id, "❌ Нет доступа.") 
            return 
        send_message(chat_id, build_kpi_text()) 
        p = build_trainplot_png() 
        send_photo(chat_id, p, caption="📈 График обучения (PNG)") 
        return 
 
    if cmd == "/trainstats": 
        if not is_admin(from_user): 
            send_message(chat_id, "❌ Команда доступна только администраторам.") 
            return 
        st = compute_training_stats() 
        last = st.get("last_ts") 
        last_s = st.get("last_str") or (datetime.fromtimestamp(last).strftime("%d.%m.%Y %H:%M") if last else "—") 
        send_message( 
            chat_id, 
            "📊 Статистика обучения:\n\n" 
            f"• Всего событий: {st['total']}\n" 
            f"   ├─ В работу: {st['work']}\n" 
            f"   └─ Неверно: {st['wrong']}\n\n" 
            f"• Прогресс к цели ({st['target']}): {st['progress']}%\n" 
            f"• Условная уверенность: {st['confidence']}%\n" 
            f"• Последнее событие: {last_s}\n" 
        ) 
        return 
 
def poll_updates_loop(): 
    global UPDATE_OFFSET 
    if not TELEGRAM_API_URL: 
        log.warning("Telegram API not configured; poller not started.") 
        return 
 
    try: 
        tg_post("deleteWebhook", {"drop_pending_updates": True}) 
    except Exception: 
        pass 
 
    log.info("Starting getUpdates poller...") 
    while True: 
        try: 
            params = {"timeout": 25, "offset": UPDATE_OFFSET, "allowed_updates": ["message","edited_message","channel_post","edited_channel_post","callback_query"]} 
            data = tg_get("getUpdates", params=params) 
            if not data: 
                time.sleep(2); continue 
 
            if not data.get("ok"):
                if data.get("error_code") == 409:
    log.error("getUpdates conflict (409). Another instance is running.")
    global LAST_CONFLICT_ALERT_TS
    # Send at most once per hour, then exit to avoid duplicate instances spamming Telegram.
    if now_ts() - LAST_CONFLICT_ALERT_TS > 3600:
        alert_msg = (
            "🚨 ВНИМАНИЕ: ОБНАРУЖЕН КОНФЛИКТ ЭКЗЕМПЛЯРОВ БОТА (ОШИБКА 409)\n\n"
            "Другой процесс/сервер уже использует этот токен Telegram, что мешает обработке обновлений.\n\n"
            "Причина: запущено несколько копий бота с одним и тем же BOT_TOKEN или включён webhook.\n"
            "Решение: оставьте только ОДИН сервис/контейнер; отключите webhook (deleteWebhook).\n\n"
            "Этот экземпляр сейчас завершится, чтобы не мешать рабочему."
        )
        recipients = list(set(list_users_by_role("admin") + list_users_by_role("leadership")))
        for uid in recipients:
            try:
                send_message(uid, alert_msg)
            except Exception:
                pass
        LAST_CONFLICT_ALERT_TS = now_ts()
    # Exit so that the "winning" instance can continue serving updates.
    raise SystemExit(0)

                log.error(f"getUpdates error: {data}")
                time.sleep(3); continue
 
            updates = data.get("result", []) or [] 
            if updates: log.info(f"[POLL] received updates={len(updates)} next_offset={UPDATE_OFFSET}") 
            if not updates: 
                continue 
 
            for upd in updates: 
                UPDATE_OFFSET = max(UPDATE_OFFSET, int(upd["update_id"]) + 1) 
                if "callback_query" in upd: 
                    handle_callback_query(upd) 
                elif any(k in upd for k in ("message","edited_message","channel_post","edited_channel_post")): 
                    handle_message(upd) 
 
            # persist offset (so restart doesn't replay) 
            set_update_offset(UPDATE_OFFSET) 
 
        except SystemExit: 
            raise 
        except Exception as e: 
            log.error(f"poll_updates exception: {e}") 
            time.sleep(3) 
 
def run_scan_cycle() -> int: 
    hits = scan_once() 
    if not hits: 
        return 0 
    sent_count = 0 
    for h in hits: 
        card = generate_card(h) 
        mid = send_card_to_group(card) 
        if mid: 
            sent_count += 1 
            time.sleep(0.4) 
    return sent_count 
 
def main(): 
    log.info("SAMASTROI SCRAPER starting...") 
    log.info(f"DATA_DIR={DATA_DIR}") 
    log.info(f"TARGET_CHAT_ID={TARGET_CHAT_ID}") 
    log.info(f"SCAN_INTERVAL={SCAN_INTERVAL}") 
    _seed_config_files() 
    log.info(f"Admins: {list_users_by_role('admin')}") 
    log.info(f"Moderators: {list_users_by_role('moderator')}") 
    log.info(f"Leadership: {list_users_by_role('leadership')}") 
    log.info(f"Prob threshold: {get_prob_threshold()}%") 
 
    acquire_lock_or_exit() 
 
    try: 
        # poller + daily reports in daemon threads 
        if str(os.getenv("ENABLE_UPDATES_POLLER", "1")).strip().lower() in ("1", "true", "yes", "on"): 
            # Important: polling conflicts with any other instance using the same BOT_TOKEN, or with an active webhook. 
            # If you need buttons/callbacks, ensure only ONE poller is running for this token. 
            try: 
                tg_post("deleteWebhook", {"drop_pending_updates": True}) 
                log.info("[POLL] deleteWebhook(drop_pending_updates=True) OK") 
            except Exception as e: 
                log.warning(f"[POLL] deleteWebhook failed: {e}") 
            log.info("[POLL] Updates poller enabled. /admin and buttons are active.") 
            threading.Thread(target=poll_updates_loop, daemon=True).start() 
        else: 
            log.info("[POLL] Updates poller disabled (set ENABLE_UPDATES_POLLER=0). Scraper-only mode.") 
        threading.Thread(target=daily_reports_worker, daemon=True).start() 
 
        while True: 
            try: 
                sent = run_scan_cycle() 
                if sent: 
                    log.info(f"Cycle done: sent={sent}") 
            except Exception as e: 
                log.error(f"scan cycle error: {e}") 
            time.sleep(SCAN_INTERVAL) 
    finally: 
        release_lock() 
 
if __name__ == "__main__": 
    main()
